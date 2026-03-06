"""
L2 Discovery Report -- VLAN documentation, routed interfaces, and findings.

Matches the example.xlsx format:
  Sheet 1: VLAN Documentation (ID, Name, Seen On, Root Bridge, Priority, SVI IPs, Notes)
  Sheet 2: Routed P2P & Loopbacks (non-SVI routed interfaces)
  Sheet 3: Legend & Findings (anomalies + MAC address map)
"""
import io
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .base import BaseReport


# ---------------------------------------------------------------------------
# Data availability helpers
# ---------------------------------------------------------------------------
def _has_collector_data(inventory_data, collector_name):
    """Check if any device has parsed data for a collector."""
    for dev in inventory_data.get("devices", {}).values():
        parsed = dev.get("collector_data", {}).get(collector_name, {}).get("parsed")
        if parsed:
            return True
    return False


def _derive_site(hostname):
    """Extract site code from hostname convention (e.g., 'NYC-CORE-01' -> 'NYC').
    Returns empty string if no pattern detected."""
    if not hostname:
        return ""
    parts = hostname.split("-")
    if len(parts) >= 2:
        return parts[0]
    return ""


# ---------------------------------------------------------------------------
# Color palette (matches example.xlsx)
# ---------------------------------------------------------------------------
TITLE_BG = "0D1B2A"
HEADER_BG = "1F3864"
WHITE = "FFFFFF"
BLACK = "000000"
ALT_ROW_BG = "F2F2F2"
ROOT_OK_BG = "E2EFDA"
ROOT_OK_FG = "375623"
ROOT_BAD_BG = "FFC7CE"
ROOT_BAD_FG = "9C0006"
MISMATCH_BG = "FFF2CC"
MISMATCH_FG = "7F6000"
NOTE_FG = "444444"
BORDER_COLOR = "D9E2EC"

# Fonts
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_DATA_FONT = Font(name="Calibri", size=11, color=BLACK)
_DATA_BOLD = Font(name="Calibri", size=11, color=BLACK, bold=True)
_NOTE_FONT = Font(name="Calibri", size=11, color=NOTE_FG)
_ROOT_OK_FONT = Font(name="Calibri", size=11, color=ROOT_OK_FG)
_ROOT_BAD_FONT = Font(name="Calibri", size=11, color=ROOT_BAD_FG, bold=True)
_MISMATCH_FONT = Font(name="Calibri", size=11, color=MISMATCH_FG)
_SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
_FINDING_TITLE_FONT = Font(name="Calibri", size=11, bold=True, color=BLACK)
_FINDING_DESC_FONT = Font(name="Calibri", size=11, color=BLACK)

# Fills
_TITLE_FILL = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
_HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
_ALT_FILL = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
_WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
_ROOT_OK_FILL = PatternFill(start_color=ROOT_OK_BG, end_color=ROOT_OK_BG, fill_type="solid")
_ROOT_BAD_FILL = PatternFill(start_color=ROOT_BAD_BG, end_color=ROOT_BAD_BG, fill_type="solid")
_MISMATCH_FILL = PatternFill(start_color=MISMATCH_BG, end_color=MISMATCH_BG, fill_type="solid")

# Alignment
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="top")
_WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)

# Border
_THIN_BORDER = Border(bottom=Side(style="thin", color=BORDER_COLOR))


# ---------------------------------------------------------------------------
# Helpers -- VLAN name matching
# ---------------------------------------------------------------------------
def _norm_vid(vlan_id):
    """Normalize 'VLAN0001' -> '1', pass through plain numbers."""
    m = re.match(r"^[Vv][Ll][Aa][Nn]0*(\d+)$", vlan_id)
    return m.group(1) if m else vlan_id


def _is_default_name(name):
    if not name:
        return True
    lower = name.strip().lower()
    if lower == "default":
        return True
    if re.match(r"^vlan0*\d+$", lower):
        return True
    return False


def _names_are_similar(a, b):
    if not a or not b:
        return True
    a_lower = a.strip().lower()
    b_lower = b.strip().lower()
    if a_lower == b_lower:
        return True
    if _is_default_name(a) and _is_default_name(b):
        return True
    if abs(len(a_lower) - len(b_lower)) > 3:
        return False
    if _levenshtein(a_lower, b_lower) <= 2:
        return True
    tokens_a = set(re.split(r"[_\-\s]+", a_lower))
    tokens_b = set(re.split(r"[_\-\s]+", b_lower))
    if tokens_a and tokens_b:
        overlap = len(tokens_a & tokens_b)
        total = max(len(tokens_a), len(tokens_b))
        if overlap / total >= 0.5:
            return True
    return False


def _levenshtein(s1, s2):
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            curr_row.append(min(prev_row[j + 1] + 1, curr_row[j] + 1,
                                prev_row[j] + (c1 != c2)))
        prev_row = curr_row
    return prev_row[-1]


# ---------------------------------------------------------------------------
# Data extraction helpers
# ---------------------------------------------------------------------------
def _extract_ip_interfaces(device):
    """Extract normalized IP interface data from a device's collector_data.

    Returns list of dicts: {interface, ips: [{ip, prefix, secondary}]}
    Handles both IOS and NX-OS ntc-templates output formats.
    """
    intf_data = device.get("collector_data", {}).get("interfaces", {})
    parsed = intf_data.get("parsed", {})
    full = parsed.get("ip_interfaces_full", [])
    results = []

    for entry in full:
        iface_name = entry.get("interface", "")
        if not iface_name:
            continue

        ips = []
        # IOS format: ip_address=[list], prefix_length=[list]
        if "ip_address" in entry and isinstance(entry["ip_address"], list):
            for idx, ip in enumerate(entry["ip_address"]):
                if not ip:
                    continue
                pfx = entry.get("prefix_length", [])
                prefix = pfx[idx] if idx < len(pfx) else ""
                ips.append({
                    "ip": ip,
                    "prefix": str(prefix),
                    "secondary": idx > 0,
                })

        # NX-OS format: primary_ip_address, primary_ip_subnet,
        #               secondary_ip_address=[list], secondary_ip_subnet=[list]
        elif "primary_ip_address" in entry:
            ip = entry.get("primary_ip_address", "")
            subnet = entry.get("primary_ip_subnet", "")
            if ip:
                # subnet is like "10.1.253.0/24", extract prefix
                prefix = subnet.split("/")[-1] if "/" in subnet else ""
                ips.append({"ip": ip, "prefix": prefix, "secondary": False})
            for idx, sec_ip in enumerate(entry.get("secondary_ip_address", [])):
                if not sec_ip:
                    continue
                sec_subnets = entry.get("secondary_ip_subnet", [])
                sec_sub = sec_subnets[idx] if idx < len(sec_subnets) else ""
                prefix = sec_sub.split("/")[-1] if "/" in sec_sub else ""
                ips.append({"ip": sec_ip, "prefix": prefix, "secondary": True})

        if ips:
            results.append({"interface": iface_name, "ips": ips})

    return results


def _get_interface_description(device, iface_name):
    """Look up interface description from interfaces collector data."""
    intf_data = device.get("collector_data", {}).get("interfaces", {})
    parsed = intf_data.get("parsed", {})
    for desc_entry in parsed.get("interfaces_description", []):
        if desc_entry.get("interface", "").lower() == iface_name.lower():
            return desc_entry.get("description", "")
    # Also check port field name variants
    for desc_entry in parsed.get("interfaces_description", []):
        port = desc_entry.get("port", desc_entry.get("interface", ""))
        if port and port.lower() == iface_name.lower():
            return desc_entry.get("description", "")
    return ""


def _get_cdp_neighbor(device, iface_name):
    """Look up CDP/LLDP neighbor for an interface."""
    cdp_data = device.get("collector_data", {}).get("cdp_lldp", {})
    parsed = cdp_data.get("parsed", {})
    for n in parsed.get("neighbors", []):
        if n.get("local_intf", "").lower() == iface_name.lower():
            return n.get("remote_device", "")
    return ""


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------
def _analyze_vlans(inventory_data):
    """Analyze VLAN data across all switches.

    Returns dict with: vlans, mismatches, name_discrepancies, blocked_ports,
                        mac_map, findings.
    """
    vlan_roots = defaultdict(dict)   # {vid: {switch: root_mac}}
    vlan_names = defaultdict(dict)   # {vid: {switch: name}}
    vlan_detail = defaultdict(dict)  # {vid: {switch: {root_address, root_priority, ...}}}
    vlan_switches = defaultdict(set) # {vid: {switch, ...}}
    blocked_ports = []

    for hostname, device in inventory_data.get("devices", {}).items():
        stp_data = device.get("collector_data", {}).get("stp_vlan", {})
        parsed = stp_data.get("parsed", {})

        for entry in parsed.get("spanning_tree_root", []):
            # Handle both normalized and raw NX-OS field names
            vid = _norm_vid(str(entry.get("vlan_id", "")))
            root_mac = entry.get("root_address", "") or entry.get("root_id", "")
            if vid and root_mac:
                vlan_roots[vid][hostname] = root_mac
                vlan_switches[vid].add(hostname)
                vlan_detail[vid][hostname] = {
                    "root_address": root_mac,
                    "root_priority": str(
                        entry.get("root_priority", "") or entry.get("priority", "")
                    ),
                    "root_cost": str(entry.get("root_cost", "")),
                    "root_port": entry.get("root_port", ""),
                }

        for entry in parsed.get("vlans", []):
            vid = _norm_vid(str(entry.get("vlan_id", "")))
            name = entry.get("name", "") or entry.get("vlan_name", "")
            if vid:
                vlan_names[vid][hostname] = name
                vlan_switches[vid].add(hostname)

        bp_entries = parsed.get("blocked_ports", [])
        # Fall back to raw parsing if parsed is empty (pre-normalization inventory)
        if not bp_entries:
            raw_stp = stp_data.get("raw", {})
            raw_blocked = raw_stp.get("show spanning-tree blockedports", "")
            raw_full = raw_stp.get("show spanning-tree", "")
            if raw_blocked.strip():
                try:
                    from app.collectors.stp_vlan import _parse_blocked_ports_raw
                except ImportError:
                    from collectors.stp_vlan import _parse_blocked_ports_raw
                bp_entries = _parse_blocked_ports_raw(raw_blocked, raw_full)

        for entry in bp_entries:
            blocked_ports.append({
                "switch": hostname,
                "vlan_id": _norm_vid(str(entry.get("vlan_id", ""))),
                "interface": entry.get("interface", ""),
                "name": entry.get("name", ""),
                "status": entry.get("status", "BLK"),
                "role": entry.get("role", ""),
                "reason": entry.get("reason", ""),
            })

    # Build MAC address map: {mac: hostname} from root entries where cost=0
    mac_map = {}
    for vid, switches in vlan_detail.items():
        for switch, detail in switches.items():
            if detail.get("root_cost") in ("0", 0):
                mac = detail.get("root_address", "")
                if mac and mac not in mac_map:
                    mac_map[mac] = switch

    # Analyze each VLAN
    vlans = {}
    mismatches = []
    name_discrepancies = []
    all_vlan_ids = sorted(
        set(list(vlan_roots.keys()) + list(vlan_names.keys())),
        key=lambda x: int(x) if x.isdigit() else 0,
    )

    for vid in all_vlan_ids:
        roots = vlan_roots.get(vid, {})
        names = vlan_names.get(vid, {})
        switches = sorted(vlan_switches.get(vid, []))

        # Unique root bridges
        root_to_switches = defaultdict(list)
        for switch, mac in roots.items():
            root_to_switches[mac].append(switch)
        unique_roots = list(root_to_switches.keys())
        consistent = len(unique_roots) <= 1

        # Determine root bridge device
        root_device = ""
        majority_root = ""
        root_priority = ""
        if unique_roots:
            majority_root = max(unique_roots, key=lambda m: len(root_to_switches[m]))
            # Resolve MAC to hostname
            root_device = mac_map.get(majority_root, "")
            # Get priority from any switch that has this root
            for switch, detail in vlan_detail.get(vid, {}).items():
                if detail.get("root_address") == majority_root:
                    root_priority = detail.get("root_priority", "")
                    break

        # Best name
        all_names = list(names.values())
        non_default = [n for n in all_names if not _is_default_name(n)]
        best_name = non_default[0] if non_default else (all_names[0] if all_names else "")

        # Name discrepancy detection
        unique_names = set(all_names)
        if len(unique_names) > 1:
            groups = []
            for name in unique_names:
                placed = False
                for group in groups:
                    if _names_are_similar(name, group[0]):
                        group.append(name)
                        placed = True
                        break
                if not placed:
                    groups.append([name])
            if len(groups) > 1:
                name_switches = defaultdict(list)
                for switch, name in names.items():
                    name_switches[name].append(switch)
                name_discrepancies.append({
                    "vlan_id": vid,
                    "names": dict(name_switches),
                })

        vlans[vid] = {
            "root_bridge_mac": majority_root,
            "root_device": root_device,
            "root_resolved": bool(root_device),
            "root_priority": root_priority,
            "root_bridges": dict(root_to_switches),
            "best_name": best_name,
            "names": dict(names),
            "switches": switches,
            "switch_count": len(switches),
            "consistent": consistent,
        }

        if not consistent:
            mismatches.append({
                "vlan_id": vid,
                "root_bridges": dict(root_to_switches),
            })

    # Build SVI IP map: {vlan_id: [{hostname, ip, prefix, secondary}]}
    svi_ips = defaultdict(list)
    for hostname, device in inventory_data.get("devices", {}).items():
        for intf in _extract_ip_interfaces(device):
            iface = intf["interface"]
            # Match Vlan interfaces
            m = re.match(r"^[Vv]lan(\d+)$", iface)
            if m:
                vid = m.group(1)
                for ip_info in intf["ips"]:
                    svi_ips[vid].append({
                        "hostname": hostname,
                        "ip": ip_info["ip"],
                        "prefix": ip_info["prefix"],
                        "secondary": ip_info["secondary"],
                    })

    # Build HSRP VIP map: {vlan_id: [{hostname, virtual_ip, state, priority, group}]}
    hsrp_vips = defaultdict(list)
    for hostname, device in inventory_data.get("devices", {}).items():
        hsrp_data = device.get("collector_data", {}).get("hsrp", {})
        parsed = hsrp_data.get("parsed", {})
        for entry in parsed.get("entries", []):
            iface = entry.get("interface", "")
            m = re.match(r"^[Vv]lan(\d+)$", iface)
            if m:
                vid = m.group(1)
                hsrp_vips[vid].append({
                    "hostname": hostname,
                    "virtual_ip": entry.get("virtual_ip", ""),
                    "state": entry.get("state", ""),
                    "priority": entry.get("priority", ""),
                    "group": entry.get("group", ""),
                })

    # Build findings
    findings = _build_findings(vlans, blocked_ports, svi_ips, inventory_data)

    return {
        "vlans": vlans,
        "svi_ips": svi_ips,
        "hsrp_vips": hsrp_vips,
        "mismatches": mismatches,
        "name_discrepancies": name_discrepancies,
        "blocked_ports": blocked_ports,
        "mac_map": mac_map,
        "findings": findings,
    }


def _build_findings(vlans, blocked_ports, svi_ips, inventory_data):
    """Auto-detect anomalies and build findings list.

    Each finding: {title, description}
    """
    findings = []

    # 1. Non-forwarding / blocked ports (any port STP isn't forwarding on)
    # Group by (switch, interface) for findings -- each unique switch+port is a finding
    by_switch_intf = defaultdict(list)
    for bp in blocked_ports:
        by_switch_intf[(bp["switch"], bp["interface"])].append(bp)

    for (switch, intf), bps in sorted(by_switch_intf.items()):
        vlan_ids = sorted(set(bp["vlan_id"] for bp in bps),
                          key=lambda x: int(x) if x.isdigit() else 0)
        statuses = sorted(set(bp["status"] for bp in bps))
        reasons = sorted(set(bp["reason"] for bp in bps if bp.get("reason")))

        status_str = "/".join(statuses)
        reason_str = f" ({', '.join(reasons)})" if reasons else ""

        vlan_str = ", ".join(vlan_ids)

        title = f"{intf} -- {status_str}{reason_str}"
        desc = (f"{intf} is {status_str}{reason_str}. "
                f"Affects VLANs: {vlan_str}.")
        findings.append({"hostname": switch, "title": title, "description": desc})

    # 2. Default STP priority on root bridges
    for vid, info in sorted(vlans.items(),
                             key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        priority_str = info.get("root_priority", "")
        if not priority_str:
            continue
        try:
            priority = int(priority_str)
        except (ValueError, TypeError):
            continue
        vlan_num = int(vid) if vid.isdigit() else 0
        # Default priority = 32768 + VLAN ID (or just 32768 for some platforms)
        if priority == 32768 + vlan_num or priority == 32768:
            root_dev = info.get("root_device") or info.get("root_bridge_mac") or "unknown"
            findings.append({
                "hostname": root_dev,
                "title": f"VLAN {vid} -- default STP priority",
                "description": (f"Root priority {priority} (32768+{vlan_num}, untuned). "
                                f"Root election is arbitrary."),
            })

    # 3. STP root bridge mismatches
    for vid, info in sorted(vlans.items(),
                             key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        if info["consistent"]:
            continue
        bridges = info["root_bridges"]
        parts = []
        for mac, switches in bridges.items():
            parts.append(f"{mac} on {', '.join(sorted(switches))}")
        findings.append({
            "hostname": "",
            "title": f"VLAN {vid} -- STP root bridge mismatch",
            "description": f"Multiple root bridges detected: {'; '.join(parts)}.",
        })

    # 4. Subnets larger than /20
    for vid, ips in sorted(svi_ips.items(),
                            key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        for ip_info in ips:
            prefix = ip_info.get("prefix", "")
            try:
                prefix_len = int(prefix)
            except (ValueError, TypeError):
                continue
            if prefix_len < 20:
                vlan_name = vlans.get(vid, {}).get("best_name", "")
                title = f"VLAN {vid}"
                if vlan_name:
                    title += f" ({vlan_name})"
                title += f" -- /{prefix_len} subnet"
                # Collect all IPs for this VLAN with this prefix
                entries = [f"{e['hostname']}: {e['ip']}/{e['prefix']}"
                           for e in ips if e.get("prefix") == prefix]
                findings.append({
                    "hostname": "",
                    "title": title,
                    "description": f"Large subnet (/{prefix_len}): {', '.join(entries)}. Verify intent.",
                })
                break  # one finding per VLAN

    # 5. VLAN name discrepancies
    for vid, info in sorted(vlans.items(),
                             key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        names = info.get("names", {})
        unique_names = set(names.values())
        if len(unique_names) <= 1:
            continue
        # Check if genuinely dissimilar
        groups = []
        for name in unique_names:
            placed = False
            for group in groups:
                if _names_are_similar(name, group[0]):
                    group.append(name)
                    placed = True
                    break
            if not placed:
                groups.append([name])
        if len(groups) > 1:
            parts = []
            name_switches = defaultdict(list)
            for switch, name in names.items():
                name_switches[name].append(switch)
            for name, switches in sorted(name_switches.items()):
                parts.append(f'"{name}" on {", ".join(sorted(switches))}')
            findings.append({
                "hostname": "",
                "title": f"VLAN {vid} -- name discrepancy",
                "description": f"Different names: {'; '.join(parts)}.",
            })

    return findings


# ---------------------------------------------------------------------------
# Worksheet builders
# ---------------------------------------------------------------------------
def _apply_cell(cell, value, font=None, fill=None, alignment=None, border=None):
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _auto_width(ws, min_width=8, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # For multi-line cells, use the longest line
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def _add_table(ws, table_name, num_cols, last_row):
    """Add an Excel table to the worksheet for auto-filter and banding."""
    if last_row < 2:
        last_row = 2  # table needs at least header + 1 row
    end_col = get_column_letter(num_cols)
    ref = f"A1:{end_col}{last_row}"
    style = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = style
    ws.add_table(table)


def _write_title_row(ws, title, num_cols, row=1):
    """Write a merged title row spanning all columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    # Fill remaining merged cells
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).fill = _TITLE_FILL


def _write_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _row_fill(row_idx, data_start):
    """Return alternating row fill."""
    if (row_idx - data_start) % 2 == 1:
        return _ALT_FILL
    return _WHITE_FILL


def _build_summary_sheet(wb, analysis, inventory_data):
    """Summary tab -- high-level metrics."""
    ws = wb.create_sheet(title="Summary")

    headers = ["Metric", "Value"]
    _write_header_row(ws, headers, row=1)

    device_count = len(inventory_data.get("devices", {}))
    vlan_count = len(analysis.get("vlans", {}))

    # Categorize findings by severity
    critical = 0
    warning = 0
    info = 0
    for finding in analysis.get("findings", []):
        sev = finding.get("severity", "")
        if sev:
            if sev == "Critical":
                critical += 1
            elif sev == "Warning":
                warning += 1
            else:
                info += 1
        else:
            # Categorize by title keywords
            title = finding.get("title", "").lower()
            if "mismatch" in title:
                critical += 1
            elif any(kw in title for kw in [
                "blk", "bkn", "default stp", "subnet",
                "bpdu", "storm", "trunk", "topology",
            ]):
                warning += 1
            else:
                info += 1

    metrics = [
        ("Device Count", device_count),
        ("VLAN Count", vlan_count),
        ("Critical Findings", critical),
        ("Warning Findings", warning),
        ("Info Findings", info),
    ]

    row = 2
    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        row += 1

    _add_table(ws, "Summary", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def _build_vlan_sheet(ws, analysis, inventory_data):
    """Sheet: VLAN Documentation."""
    headers = ["VLAN ID", "Site/Location", "VLAN Name", "Seen On (Switches)",
               "STP Root Bridge", "Root Priority", "SVI IP / CIDR",
               "HSRP VIP", "Blocked Ports", "Notes"]
    num_cols = len(headers)

    _write_header_row(ws, headers, row=1)

    # Pre-build per-VLAN blocked port summary
    bp_by_vlan = defaultdict(list)
    bp_by_switch_intf = defaultdict(set)
    for bp in analysis.get("blocked_ports", []):
        bp_by_switch_intf[(bp["switch"], bp["interface"])].add(bp["vlan_id"])
    for (switch, intf), vlan_set in bp_by_switch_intf.items():
        sorted_vlans = sorted(vlan_set, key=lambda x: int(x) if x.isdigit() else 0)
        for vid in sorted_vlans:
            bp_by_vlan[vid].append({
                "switch": switch,
                "interface": intf,
                "all_vlans": sorted_vlans,
            })

    row = 2
    for vid, info in sorted(analysis["vlans"].items(),
                             key=lambda x: int(x[0]) if x[0].isdigit() else 0):

        # Col A: VLAN ID
        ws.cell(row=row, column=1, value=int(vid) if vid.isdigit() else vid)

        # Col B: Site/Location
        sites = sorted(set(
            _derive_site(s) for s in info["switches"] if _derive_site(s)
        ))
        ws.cell(row=row, column=2, value=", ".join(sites))

        # Col C: VLAN Name
        ws.cell(row=row, column=3, value=info["best_name"])

        # Col D: Seen On (Switches)
        ws.cell(row=row, column=4, value=", ".join(sorted(info["switches"])))

        # Col E: STP Root Bridge (device name or MAC with conditional fill)
        if info["root_resolved"]:
            cell = ws.cell(row=row, column=5, value=info["root_device"])
            cell.font = _ROOT_OK_FONT
            cell.fill = _ROOT_OK_FILL
        elif info["root_bridge_mac"]:
            cell = ws.cell(row=row, column=5, value=info["root_bridge_mac"])
            cell.font = _ROOT_BAD_FONT
            cell.fill = _ROOT_BAD_FILL
        else:
            ws.cell(row=row, column=5, value="")

        # Col F: Root Priority
        priority_val = info["root_priority"]
        if priority_val:
            try:
                priority_val = int(priority_val)
            except (ValueError, TypeError):
                pass
        ws.cell(row=row, column=6, value=priority_val if priority_val else "")

        # Col G: SVI IP / CIDR
        svi_entries = analysis["svi_ips"].get(vid, [])
        if svi_entries:
            by_host = defaultdict(list)
            for e in svi_entries:
                by_host[e["hostname"]].append(e)
            lines = []
            for hostname in sorted(by_host.keys()):
                for e in by_host[hostname]:
                    cidr = f"{e['ip']}/{e['prefix']}" if e["prefix"] else e["ip"]
                    label = f"{hostname}: {cidr}"
                    if e["secondary"]:
                        label += " (secondary)"
                    lines.append(label)
            svi_text = "\n".join(lines)
        else:
            svi_text = ""
        cell = ws.cell(row=row, column=7, value=svi_text)
        if "\n" in svi_text:
            cell.alignment = _WRAP_ALIGN

        # Col H: HSRP VIP
        hsrp_entries = analysis.get("hsrp_vips", {}).get(vid, [])
        if hsrp_entries:
            lines = []
            for e in sorted(hsrp_entries, key=lambda x: x["hostname"]):
                vip = e["virtual_ip"]
                state = e.get("state", "")
                prio = e.get("priority", "")
                label = f"{e['hostname']}: {vip}"
                parts = []
                if state:
                    parts.append(state)
                if prio:
                    parts.append(f"pri {prio}")
                if parts:
                    label += f" ({', '.join(parts)})"
                lines.append(label)
            hsrp_text = "\n".join(lines)
        else:
            hsrp_text = ""
        cell = ws.cell(row=row, column=8, value=hsrp_text)
        if "\n" in hsrp_text:
            cell.alignment = _WRAP_ALIGN

        # Col I: Blocked Ports
        bp_entries_for_vlan = bp_by_vlan.get(vid, [])
        if bp_entries_for_vlan:
            lines = []
            for bp in sorted(bp_entries_for_vlan, key=lambda x: (x["switch"], x["interface"])):
                all_vlans = bp["all_vlans"]
                line = f"{bp['switch']}-{bp['interface']} BLK on VLANs {', '.join(all_vlans)}"
                lines.append(line)
            bp_text = "\n".join(lines)
        else:
            bp_text = ""
        cell = ws.cell(row=row, column=9, value=bp_text)
        if "\n" in bp_text:
            cell.alignment = _WRAP_ALIGN

        # Col J: Notes
        notes = []
        names = info.get("names", {})
        unique_names = set(names.values())
        if len(unique_names) > 1:
            groups = []
            for name in unique_names:
                placed = False
                for group in groups:
                    if _names_are_similar(name, group[0]):
                        group.append(name)
                        placed = True
                        break
                    if not placed:
                        groups.append([name])
            if len(groups) > 1:
                notes.append(f"Name varies: {', '.join(sorted(unique_names))}")
        if not info["consistent"]:
            notes.append("STP root mismatch -- see Findings")
        ws.cell(row=row, column=10, value="; ".join(notes) if notes else "")

        row += 1

    _add_table(ws, "VLANDocumentation", num_cols, row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_findings_sheet(ws, analysis):
    """Findings sheet -- Hostname, Site/Location, Port, Notes."""
    findings = analysis["findings"]

    headers = ["Hostname", "Site/Location", "Port", "Notes"]
    _write_header_row(ws, headers, row=1)

    row = 2
    if findings:
        for finding in findings:
            hostname = finding.get("hostname", "")
            ws.cell(row=row, column=1, value=hostname)
            ws.cell(row=row, column=2, value=_derive_site(hostname))
            ws.cell(row=row, column=3, value=finding["title"])
            cell = ws.cell(row=row, column=4, value=finding["description"])
            cell.alignment = _WRAP_ALIGN
            row += 1
    else:
        ws.cell(row=row, column=1, value="No findings detected.")
        row += 1

    _add_table(ws, "Findings", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_root_bridge_sheet(ws, analysis):
    """Root Bridge IDs -- Hostname, Site/Location, base MAC mapping."""
    mac_map = analysis["mac_map"]

    headers = ["Hostname", "Site/Location", "Root Bridge ID"]
    _write_header_row(ws, headers, row=1)

    host_to_mac = {}
    for mac, hostname in mac_map.items():
        if hostname not in host_to_mac:
            host_to_mac[hostname] = mac

    row = 2
    for hostname in sorted(host_to_mac.keys()):
        ws.cell(row=row, column=1, value=hostname)
        ws.cell(row=row, column=2, value=_derive_site(hostname))
        ws.cell(row=row, column=3, value=host_to_mac[hostname])
        row += 1

    _add_table(ws, "RootBridgeIDs", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_stp_topology_sheet(wb, inventory_data):
    """STP Topology tab -- per-VLAN STP detail per device."""
    ws = wb.create_sheet(title="STP Topology")

    headers = ["Device", "Site/Location", "VLAN", "Interface", "Role",
               "State", "Cost", "Topology Changes"]
    _write_header_row(ws, headers, row=1)

    row = 2
    for hostname in sorted(inventory_data.get("devices", {}).keys()):
        device = inventory_data["devices"][hostname]
        stp_detail = device.get("collector_data", {}).get("stp_detail", {})
        parsed = stp_detail.get("parsed", {})
        entries = parsed.get("stp_detail", [])

        # Re-parse from raw if parsed data is empty (pre-fix inventories)
        if not entries:
            raw = stp_detail.get("raw", {})
            raw_detail = raw.get("show spanning-tree detail", "")
            if raw_detail:
                try:
                    from app.collectors.stp_detail import _parse_stp_detail
                except ImportError:
                    from collectors.stp_detail import _parse_stp_detail
                entries = _parse_stp_detail(raw_detail)

        site = _derive_site(hostname)

        for entry in entries:
            topo_changes = 0
            try:
                topo_changes = int(entry.get("topology_changes", 0))
            except (ValueError, TypeError):
                topo_changes = 0

            ws.cell(row=row, column=1, value=hostname)
            ws.cell(row=row, column=2, value=site)
            ws.cell(row=row, column=3, value=entry.get("vlan", ""))
            ws.cell(row=row, column=4, value=entry.get("interface", ""))
            ws.cell(row=row, column=5, value=entry.get("role", ""))
            ws.cell(row=row, column=6, value=entry.get("state", ""))
            ws.cell(row=row, column=7, value=entry.get("cost", ""))
            ws.cell(row=row, column=8, value=topo_changes)
            row += 1

    _add_table(ws, "STPTopology", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def _build_port_security_sheet(wb, inventory_data):
    """Port Security tab -- access port hardening analysis."""
    ws = wb.create_sheet(title="Port Security")

    headers = ["Device", "Site/Location", "Interface", "Mode", "BPDU Guard",
               "Root Guard", "Storm Control", "Port Security", "Voice VLAN"]
    _write_header_row(ws, headers, row=1)

    row = 2
    for hostname in sorted(inventory_data.get("devices", {}).keys()):
        device = inventory_data["devices"][hostname]
        sw_data = device.get("collector_data", {}).get("switchport", {})
        parsed = sw_data.get("parsed", {})
        site = _derive_site(hostname)

        # Get running config for BPDU/root guard detection
        config_text = device.get("collector_data", {}).get(
            "config", {}).get("parsed", {}).get("config", "")
        global_bpdu = ("spanning-tree portfast bpduguard default" in config_text)

        # Build storm control lookup by interface
        storm_lookup = {}
        for sc in parsed.get("storm_control", []):
            iface = sc.get("interface", "")
            if iface:
                storm_lookup[iface.lower()] = sc

        # Build port security lookup by interface
        ps_lookup = {}
        for ps in parsed.get("port_security", []):
            iface = ps.get("interface", "")
            if iface:
                ps_lookup[iface.lower()] = ps

        for port in parsed.get("switchports", []):
            iface = port.get("interface", "")
            mode = port.get("mode", "")

            # BPDU Guard: global or per-interface in config
            has_bpdu = global_bpdu
            if not has_bpdu and config_text and iface:
                if "spanning-tree bpduguard enable" in config_text:
                    has_bpdu = True
            bpdu_str = "Yes" if has_bpdu else "No"

            # Root Guard
            has_root_guard = False
            if config_text:
                if "spanning-tree guard root" in config_text:
                    has_root_guard = True
            root_guard_str = "Yes" if has_root_guard else "No"

            # Storm Control
            has_storm = iface.lower() in storm_lookup
            storm_str = "Yes" if has_storm else "No"

            # Port Security
            has_ps = iface.lower() in ps_lookup
            ps_str = "Yes" if has_ps else "No"

            # Voice VLAN
            voice_vlan = port.get("voice_vlan", "")

            ws.cell(row=row, column=1, value=hostname)
            ws.cell(row=row, column=2, value=site)
            ws.cell(row=row, column=3, value=iface)
            ws.cell(row=row, column=4, value=mode)
            ws.cell(row=row, column=5, value=bpdu_str)
            ws.cell(row=row, column=6, value=root_guard_str)
            ws.cell(row=row, column=7, value=storm_str)
            ws.cell(row=row, column=8, value=ps_str)
            ws.cell(row=row, column=9, value=voice_vlan)

            # Highlight missing security features based on port mode:
            #   ACCESS: BPDU Guard (5), Storm Control (7)
            #   TRUNK:  Root Guard (6), Storm Control (7)
            is_trunk = "trunk" in mode.lower()
            is_access = "access" in mode.lower() or "static access" in mode.lower()
            highlight_cols = []
            if is_access:
                highlight_cols = [5, 7]   # BPDU Guard, Storm Control
            elif is_trunk:
                highlight_cols = [6, 7]   # Root Guard, Storm Control
            for col in highlight_cols:
                if ws.cell(row=row, column=col).value == "No":
                    ws.cell(row=row, column=col).fill = _ROOT_BAD_FILL

            row += 1

    _add_table(ws, "PortSecurity", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def _build_trunk_summary_sheet(wb, inventory_data):
    """Trunk Summary tab -- trunk port analysis."""
    ws = wb.create_sheet(title="Trunk Summary")

    headers = ["Device", "Site/Location", "Interface", "Native VLAN",
               "Allowed VLANs", "Neighbor"]
    _write_header_row(ws, headers, row=1)

    row = 2
    for hostname in sorted(inventory_data.get("devices", {}).keys()):
        device = inventory_data["devices"][hostname]
        sw_data = device.get("collector_data", {}).get("switchport", {})
        parsed = sw_data.get("parsed", {})
        site = _derive_site(hostname)

        for port in parsed.get("switchports", []):
            mode = port.get("mode", "")
            if "trunk" not in mode.lower():
                continue

            iface = port.get("interface", "")
            native_vlan = port.get("native_vlan", "")
            allowed_vlans = port.get("allowed_vlans", "")

            # Get neighbor from cdp_lldp
            neighbor = _get_cdp_neighbor(device, iface)

            ws.cell(row=row, column=1, value=hostname)
            ws.cell(row=row, column=2, value=site)
            ws.cell(row=row, column=3, value=iface)
            ws.cell(row=row, column=4, value=native_vlan)
            ws.cell(row=row, column=5, value=allowed_vlans)
            ws.cell(row=row, column=6, value=neighbor)

            # Highlight VLAN 1 native with light red
            if str(native_vlan) == "1":
                ws.cell(row=row, column=4).fill = _ROOT_BAD_FILL

            # Highlight unpruned trunks with light red
            if allowed_vlans in ("ALL", "1-4094", ""):
                ws.cell(row=row, column=5).fill = _ROOT_BAD_FILL

            row += 1

    _add_table(ws, "TrunkSummary", len(headers), row - 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# Report class
# ---------------------------------------------------------------------------
class L2DiscoveryReport(BaseReport):
    name = "l2_discovery"
    label = "L2 Discovery Report"
    description = "VLAN documentation, routed interfaces, and network findings"
    category = "Layer 2 Analysis"
    required_collectors = ["stp_vlan", "cdp_lldp", "interfaces"]
    supported_formats = ["xlsx"]

    def generate(self, inventory_data, fmt="xlsx"):
        analysis = _analyze_vlans(inventory_data)

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Summary (always)
        _build_summary_sheet(wb, analysis, inventory_data)

        # Sheet 2: VLAN Documentation (always)
        ws2 = wb.create_sheet(title="VLAN Documentation")
        _build_vlan_sheet(ws2, analysis, inventory_data)

        # Sheet 3: Findings (always)
        ws3 = wb.create_sheet(title="Findings")
        _build_findings_sheet(ws3, analysis)

        # Sheet 4: Root Bridge IDs (always)
        ws4 = wb.create_sheet(title="Root Bridge IDs")
        _build_root_bridge_sheet(ws4, analysis)

        # Sheet 5: STP Topology (if stp_detail data exists)
        if _has_collector_data(inventory_data, "stp_detail"):
            _build_stp_topology_sheet(wb, inventory_data)

        # Sheet 6: Port Security (if switchport data exists)
        if _has_collector_data(inventory_data, "switchport"):
            _build_port_security_sheet(wb, inventory_data)

        # Sheet 7: Trunk Summary (if switchport data exists)
        if _has_collector_data(inventory_data, "switchport"):
            _build_trunk_summary_sheet(wb, inventory_data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
