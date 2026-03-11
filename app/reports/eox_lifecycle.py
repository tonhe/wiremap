"""
End-of-Life / End-of-Support report.
Looks up Cisco hardware PIDs and software versions against the EOX API,
produces a 4-tab XLSX: Summary, Hardware Lifecycle, Software Lifecycle, Modules & Components.
"""

import io
import json
import logging
import os
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from .base import BaseReport
from .xlsx_utils import create_sheet, style_header_row, write_data_rows, auto_width, freeze_header, add_table

logger = logging.getLogger(__name__)

# Status thresholds
APPROACHING_DAYS = 365  # "Approaching EoS" if LDoS is within this many days

# Status labels
STATUS_EOS = "End-of-Support"
STATUS_APPROACHING = "Approaching EoS"
STATUS_END_OF_SALE = "End-of-Sale"
STATUS_CURRENT = "Current"
STATUS_NOT_IN_EOX = "Not in EOX"       # PID looked up, API has no record for it
STATUS_NO_PID = "No Platform ID"       # No PID available to look up
STATUS_NO_DATA = "No Data"             # Fallback / unexpected

# Status fill colors for XLSX
STATUS_FILLS = {
    STATUS_EOS: PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid"),          # Red
    STATUS_APPROACHING: PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid"),   # Amber
    STATUS_END_OF_SALE: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),   # Yellow
    STATUS_CURRENT: PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid"),       # Green
    STATUS_NOT_IN_EOX: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),    # Gray
    STATUS_NO_PID: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),        # Gray
    STATUS_NO_DATA: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),       # Gray
}

# Cache TTL
CACHE_TTL_DAYS = 7

# Patterns for chassis/switch PIDs (not modules)
_CHASSIS_MODEL_RE = re.compile(
    r"^("
    r"WS-C\d|"
    r"C9[2-6]\d{2}[LX]?-\d|"
    r"CISCO\d|"
    r"ISR\d.*/K9|"
    r"VG[23]\d{2}|"
    r"ASR\d|"
    r"N[579]K-C\d|"
    r"NCS-\d"
    r")", re.IGNORECASE
)


def _normalize_keys(d):
    return {k.lower(): v for k, v in d.items()}


def _parse_date(date_str):
    """Parse YYYY-MM-DD string to date object, or None."""
    if not date_str or not date_str.strip():
        return None
    try:
        return datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _compute_status(eox_record):
    """Determine lifecycle status from an EOX record."""
    if not eox_record:
        return STATUS_NO_DATA, None
    if eox_record.get("_pid_not_found"):
        return STATUS_NOT_IN_EOX, None
    if eox_record.get("_no_pid"):
        return STATUS_NO_PID, None

    ldos_str = eox_record.get("last_date_of_support", "")
    eos_str = eox_record.get("end_of_sale", "")
    ldos = _parse_date(ldos_str)
    eos = _parse_date(eos_str)
    today = date.today()

    if ldos:
        days_remaining = (ldos - today).days
        if days_remaining < 0:
            return STATUS_EOS, days_remaining
        elif days_remaining <= APPROACHING_DAYS:
            return STATUS_APPROACHING, days_remaining
        elif eos and eos <= today:
            return STATUS_END_OF_SALE, days_remaining
        else:
            return STATUS_CURRENT, days_remaining
    elif eos:
        if eos <= today:
            return STATUS_END_OF_SALE, None
        else:
            return STATUS_CURRENT, None
    else:
        # No dates at all in record
        return STATUS_CURRENT, None


_SLOT_NAME_RE = re.compile(r'^\s*(module\s+)?slot\s+\d', re.IGNORECASE)
_CONSUMABLE_RE = re.compile(r'(^|[\b_\-])(fan|power|psu|sfp|transceiver|xcvr|cwdm|dwdm|gbic)($|[\b_\-])', re.IGNORECASE)

# ISR built-in interface designations like ISR4331-3x1GE — not separately orderable PIDs
_ISR_INLINE_RE = re.compile(r'^ISR\d+[A-Z0-9\-]*-\d+x\d+GE$', re.IGNORECASE)


def _is_consumable(name, descr="", pid=""):
    """Return True if this inventory item should be excluded from the module tab."""
    if _ISR_INLINE_RE.match(pid or ""):
        return True
    return bool(
        _CONSUMABLE_RE.search(name or "")
        or _CONSUMABLE_RE.search(descr or "")
        or _CONSUMABLE_RE.search(pid or "")
    )

# MAC address patterns (Cisco dotted and standard colon formats)
_MAC_RE = re.compile(r'^([0-9a-f]{4}\.){2}[0-9a-f]{4}$|^([0-9a-f]{2}:){5}[0-9a-f]{2}$', re.IGNORECASE)

# IP address pattern
_IP_RE = re.compile(r'^\d{1,3}(\.\d{1,3}){3}$')

# Leading "cisco " / "Cisco " vendor prefix in platform strings from CDP sysDescr
_CISCO_PREFIX_RE = re.compile(r'^cisco\s+', re.IGNORECASE)

# CDP platform strings often embed the real PID: "SG350-28P (PID:SG350-28P-K9)-VSD"
_CDP_PID_RE = re.compile(r'\(PID:([^)]+)\)', re.IGNORECASE)

# NX-OS versions have a trailing sub-release like 7.0(3)I7(5) — strip the last (N)
_NXOS_SUBRELEASE_RE = re.compile(r'^(\d+\.\d+\(\d+\)\w+)\(\d+\)$')

# IOS minor release with trailing letter: 15.0(2a)EX5 -> 15.0(2)EX5
_IOS_MINOR_LETTER_RE = re.compile(r'\((\d+)[a-z]\)', re.IGNORECASE)

# IOS release train with trailing letter after number: 15.5(3)S4b -> 15.5(3)S4
_IOS_TRAIN_LETTER_RE = re.compile(r'([A-Z]+\d+)[a-z]$')


def _normalize_sw_version(version, os_type):
    """Normalize a software version string for EOX API lookup."""
    if not version:
        return version
    if os_type == "NX-OS":
        # Strip trailing sub-release: 7.0(3)I7(5) -> 7.0(3)I7
        m = _NXOS_SUBRELEASE_RE.match(version)
        return m.group(1) if m else version
    # IOS/IOS-XE/IOS-XR: strip letter from minor release and release train
    version = _IOS_MINOR_LETTER_RE.sub(r'(\1)', version)
    version = _IOS_TRAIN_LETTER_RE.sub(r'\1', version)
    return version


def _clean_pid(pid):
    """Normalize a PID string from CDP/NTC data into a clean EOX-lookupable PID."""
    if not pid:
        return ""
    pid = pid.strip()
    # Reject non-printable / control characters (e.g. 0x7F DEL string from bad NTC parse)
    if any(ord(c) < 0x20 or ord(c) == 0x7F for c in pid):
        return ""
    if _MAC_RE.match(pid) or _IP_RE.match(pid):
        return ""
    # Extract embedded PID from CDP platform strings like "SG350-28P (PID:SG350-28P-K9)-VSD"
    m = _CDP_PID_RE.search(pid)
    if m:
        return m.group(1).strip()
    # Strip leading "cisco " vendor prefix from CDP sysDescr
    pid = _CISCO_PREFIX_RE.sub("", pid).strip()
    return pid


def _is_chassis_pid(pid):
    return bool(_CHASSIS_MODEL_RE.match(pid)) if pid else False


def _is_slot_name(name):
    """Return True if an inventory item name indicates a module slot (covered by show module)."""
    return bool(_SLOT_NAME_RE.match(name)) if name else False


class EoxLifecycleReport(BaseReport):
    name = "eox_lifecycle"
    label = "End-of-Life / End-of-Support"
    description = "Cisco hardware and software lifecycle dates from the EOX API"
    category = "Compliance & Config"
    required_collectors = ["device_inventory"]
    supported_formats = ["xlsx"]

    def can_generate(self, inventory_data):
        """Override: also check that the EOX plugin is configured."""
        if not super().can_generate(inventory_data):
            return False
        # Check plugin availability
        try:
            from plugins import get_plugin_config
            cfg = get_plugin_config("cisco_eox")
            return cfg is not None and cfg.get("enabled") and cfg.get("client_id")
        except ImportError:
            return False

    def generate(self, inventory_data, fmt="xlsx"):
        # Load EOX data (from cache or API)
        hw_eox, sw_eox = self._get_eox_data(inventory_data)

        wb = Workbook()
        wb.remove(wb.active)

        hw_rows = self._build_hardware_rows(inventory_data, hw_eox)
        sw_rows = self._build_software_rows(inventory_data, sw_eox)
        mod_rows = self._build_module_rows(inventory_data, hw_eox)

        self._build_summary_sheet(wb, inventory_data, hw_rows, sw_rows, mod_rows, hw_eox, sw_eox)
        self._build_lifecycle_sheet(wb, "Hardware Lifecycle", hw_rows)
        self._build_lifecycle_sheet(wb, "Software Lifecycle", sw_rows, is_software=True)
        self._build_lifecycle_sheet(wb, "Modules & Components", mod_rows)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _get_eox_data(self, inventory_data):
        """Get EOX data from cache or API. Returns (hw_eox_dict, sw_eox_dict)."""
        # Extract all unique PIDs and SW versions
        pids = set()
        sw_pairs = set()

        for hostname, device in inventory_data.get("devices", {}).items():
            if not device.get("device_type", "").startswith("cisco"):
                continue

            inv_data = device.get("collector_data", {}).get("device_inventory", {})
            parsed = inv_data.get("parsed", {})

            # Hardware PIDs — from show inventory, stack members, modules
            found_chassis_pid = False
            for item in parsed.get("inventory", []):
                it = _normalize_keys(item)
                pid = _clean_pid(it.get("pid") or "")
                if pid:
                    pids.add(pid)
                    name_lower = (it.get("name") or "").lower().strip()
                    if name_lower in ("chassis", "") or re.match(r'^switch\s+1$', name_lower):
                        found_chassis_pid = True

            for member in parsed.get("stack_members", []):
                m = _normalize_keys(member)
                model = _clean_pid(m.get("model") or "")
                if model:
                    pids.add(model)
                    found_chassis_pid = True

            for mod in parsed.get("modules", []):
                m = _normalize_keys(mod)
                model = _clean_pid(m.get("model") or "")
                if model:
                    pids.add(model)

            # Always include device.platform as a candidate PID — it's often more
            # accurate than show version hardware (e.g. 'N7K-C7009' vs 'C7009').
            # If inventory already found a chassis PID, platform is still useful
            # as a fallback lookup in case the primary PID isn't in EOX.
            platform = _clean_pid(device.get("platform") or "")
            if platform:
                pids.add(platform)
            elif not found_chassis_pid:
                pass  # no platform and no inventory — chassis will show No Platform ID

            # Software versions
            from eox_client import EoxClient
            os_type = EoxClient.get_os_type(device.get("device_type", ""))
            if os_type:
                version_list = parsed.get("version", [])
                if version_list and isinstance(version_list, list):
                    v = _normalize_keys(version_list[0])
                    sw_ver = v.get("version") or v.get("os_version") or v.get("os") or ""
                    if isinstance(sw_ver, list):
                        sw_ver = sw_ver[0] if sw_ver else ""
                    if sw_ver:
                        sw_ver = _normalize_sw_version(sw_ver, os_type)
                        sw_pairs.add((sw_ver, os_type))

        # Check cache
        eox_cache = inventory_data.get("eox_cache", {})
        today = date.today()

        uncached_pids = []
        hw_eox = {}
        for pid in pids:
            cache_key = f"hw:{pid}"
            cached = eox_cache.get(cache_key)
            if cached and self._cache_valid(cached, today):
                hw_eox[pid] = cached["record"]
            else:
                uncached_pids.append(pid)

        uncached_sw = []
        sw_eox = {}
        for ver, os_type in sw_pairs:
            cache_key = f"sw:{ver}|{os_type}"
            cached = eox_cache.get(cache_key)
            if cached and self._cache_valid(cached, today):
                sw_eox[f"{ver}|{os_type}"] = cached["record"]
            else:
                uncached_sw.append((ver, os_type))

        # Fetch from API if needed
        if uncached_pids or uncached_sw:
            try:
                from plugins import get_plugin_config
                from eox_client import EoxClient

                cfg = get_plugin_config("cisco_eox")
                if cfg and cfg.get("client_id") and cfg.get("client_secret"):
                    client = EoxClient(cfg["client_id"], cfg["client_secret"])

                    if uncached_pids:
                        logger.info("Looking up %d PIDs from EOX API", len(uncached_pids))
                        pid_results = client.lookup_pids(uncached_pids)
                        for pid, record in pid_results.items():
                            hw_eox[pid] = record
                            eox_cache[f"hw:{pid}"] = {
                                "fetched_at": datetime.utcnow().isoformat() + "Z",
                                "ttl_days": CACHE_TTL_DAYS,
                                "record": record,
                            }

                    if uncached_sw:
                        logger.info("Looking up %d SW versions from EOX API", len(uncached_sw))
                        sw_results = client.lookup_software(uncached_sw)
                        for key, record in sw_results.items():
                            sw_eox[key] = record
                            eox_cache[f"sw:{key}"] = {
                                "fetched_at": datetime.utcnow().isoformat() + "Z",
                                "ttl_days": CACHE_TTL_DAYS,
                                "record": record,
                            }

                    # Persist cache back to inventory file
                    self._save_cache(inventory_data, eox_cache)

            except Exception as e:
                logger.error("EOX API lookup failed: %s", e)

        return hw_eox, sw_eox

    def _cache_valid(self, cached, today):
        fetched = cached.get("fetched_at", "")
        ttl = cached.get("ttl_days", CACHE_TTL_DAYS)
        if not fetched:
            return False
        try:
            fetched_date = datetime.fromisoformat(fetched.rstrip("Z")).date()
            return (today - fetched_date).days < ttl
        except ValueError:
            return False

    def _save_cache(self, inventory_data, eox_cache):
        """Persist the EOX cache back to the inventory JSON file."""
        discovery_id = inventory_data.get("discovery_id", "")
        if not discovery_id:
            return
        inventory_dir = "/app/inventories"
        filepath = os.path.join(inventory_dir, f"{discovery_id}.json")
        if not os.path.exists(filepath):
            return
        try:
            with open(filepath, "r") as f:
                saved = json.load(f)
            saved["eox_cache"] = eox_cache
            with open(filepath, "w") as f:
                json.dump(saved, f)
        except Exception as e:
            logger.warning("Failed to save EOX cache: %s", e)

    def _build_hardware_rows(self, inventory_data, hw_eox):
        """Build rows for Hardware Lifecycle tab. One row per device chassis + stack members."""
        rows = []
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            if not device.get("device_type", "").startswith("cisco"):
                continue

            mgmt_ip = device.get("mgmt_ip", "")
            inv_data = device.get("collector_data", {}).get("device_inventory", {})
            parsed = inv_data.get("parsed", {})

            # Chassis PID from inventory or version
            chassis_pid = ""
            chassis_serial = ""
            chassis_descr = ""

            version_list = parsed.get("version", [])
            if version_list and isinstance(version_list, list):
                v = _normalize_keys(version_list[0])
                chassis_pid = v.get("hardware") or v.get("platform") or ""
                if isinstance(chassis_pid, list):
                    chassis_pid = chassis_pid[0] if chassis_pid else ""
                chassis_pid = _clean_pid(chassis_pid)
                chassis_serial = v.get("serial") or v.get("serial_number") or ""
                if isinstance(chassis_serial, list):
                    chassis_serial = chassis_serial[0] if chassis_serial else ""

            # Fallback to device-level platform (CDP-discovered devices with no collector data)
            if not chassis_pid:
                chassis_pid = _clean_pid(device.get("platform") or "")

            # Try to get better PID from show inventory (first chassis entry)
            inv_list = parsed.get("inventory", [])
            switch1_inv = {}  # fallback for stacked switches with no explicit chassis entry
            for item in inv_list:
                it = _normalize_keys(item)
                name = (it.get("name") or "").strip()
                name_lower = name.lower()
                if "chassis" in name_lower or name_lower == "1":
                    pid = _clean_pid(it.get("pid") or "")
                    if pid:
                        chassis_pid = pid
                    sn = it.get("sn") or it.get("serial_number") or ""
                    if sn:
                        chassis_serial = sn
                    chassis_descr = it.get("descr") or it.get("description") or ""
                    break
                # Capture Switch 1 inventory entry as fallback chassis PID for stacks
                sw_match = re.match(r'^[Ss]witch\s+1$', name)
                if sw_match and not switch1_inv:
                    pid = _clean_pid(it.get("pid") or "")
                    if pid:
                        switch1_inv = {
                            "pid": pid,
                            "sn": it.get("sn") or it.get("serial_number") or "",
                            "descr": it.get("descr") or it.get("description") or "",
                        }

            # For stacked switches: if no explicit chassis PID was found via inventory,
            # use Switch 1 inventory entry (it IS the chassis for member 1)
            if not chassis_pid and switch1_inv:
                chassis_pid = switch1_inv["pid"]
                if not chassis_serial:
                    chassis_serial = switch1_inv["sn"]
                if not chassis_descr:
                    chassis_descr = switch1_inv["descr"]

            if chassis_pid:
                eox = hw_eox.get(chassis_pid)
                # If the primary PID wasn't found in EOX, try device.platform as a
                # better alternative (e.g. show version returns 'C7009' but
                # device.platform is 'N7K-C7009' which IS in the EOX API).
                if (not eox or eox.get("_pid_not_found")):
                    alt_pid = _clean_pid(device.get("platform") or "")
                    if alt_pid and alt_pid != chassis_pid:
                        alt_eox = hw_eox.get(alt_pid)
                        if alt_eox and not alt_eox.get("_pid_not_found"):
                            chassis_pid = alt_pid
                            eox = alt_eox
            else:
                eox = {"_no_pid": True}
            status, days = _compute_status(eox)

            rows.append(self._make_hw_row(
                hostname, mgmt_ip, "Chassis", chassis_pid,
                eox.get("description", "") if eox and not eox.get("_no_pid") else chassis_descr,
                chassis_serial, eox, status, days,
            ))

            # Build per-member PID/serial lookup from show inventory
            # Stack member entries are named "Switch N" or "Switch N System"
            member_inv = {}  # "1", "2", ... -> {"pid": ..., "sn": ..., "descr": ...}
            for item in parsed.get("inventory", []):
                it = _normalize_keys(item)
                name = (it.get("name") or "").strip()
                sw_match = re.match(r'^[Ss]witch\s+(\d+)', name)
                if sw_match:
                    num = sw_match.group(1)
                    pid = _clean_pid(it.get("pid") or "")
                    if pid and num not in member_inv:
                        member_inv[num] = {
                            "pid": pid,
                            "sn": it.get("sn") or it.get("serial_number") or "",
                            "descr": it.get("descr") or it.get("description") or "",
                        }

            # Stack members — prefer per-member PID from inventory for mixed stacks
            for member in parsed.get("stack_members", []):
                m = _normalize_keys(member)
                switch_num = str(m.get("switch") or m.get("member") or "")
                inv_entry = member_inv.get(switch_num, {})
                model = inv_entry.get("pid") or _clean_pid(m.get("model") or "") or chassis_pid
                serial = inv_entry.get("sn") or m.get("serial") or m.get("serialnum") or ""
                eox = hw_eox.get(model)
                status, days = _compute_status(eox)
                rows.append(self._make_hw_row(
                    hostname, mgmt_ip, f"Stack Member {switch_num}",
                    model, eox.get("description", "") if eox else inv_entry.get("descr", ""),
                    serial, eox, status, days,
                ))

        # Sort by days remaining (most urgent first), blanks/None at end
        rows.sort(key=lambda r: (r[11] == "" or r[11] is None, r[11] if isinstance(r[11], int) else 999999))
        return rows

    def _build_module_rows(self, inventory_data, hw_eox):
        """Build rows for Modules & Components tab."""
        rows = []
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            if not device.get("device_type", "").startswith("cisco"):
                continue

            mgmt_ip = device.get("mgmt_ip", "")
            inv_data = device.get("collector_data", {}).get("device_inventory", {})
            parsed = inv_data.get("parsed", {})

            # From show module
            for mod in parsed.get("modules", []):
                m = _normalize_keys(mod)
                model = m.get("model") or ""
                if not model or _is_chassis_pid(model):
                    continue
                if _is_consumable(model):
                    continue
                serial = m.get("serial") or m.get("serialnum") or ""
                slot = m.get("module") or m.get("slot") or ""
                eox = hw_eox.get(model)
                status, days = _compute_status(eox)
                rows.append(self._make_hw_row(
                    hostname, mgmt_ip, f"Module Slot {slot}",
                    model, eox.get("description", "") if eox else "",
                    serial, eox, status, days,
                ))

            # From show inventory (non-chassis, non-module items)
            for item in parsed.get("inventory", []):
                it = _normalize_keys(item)
                name = it.get("name") or ""
                pid = it.get("pid") or ""
                if not pid or "chassis" in name.lower() or _is_slot_name(name):
                    continue
                if _is_chassis_pid(pid):
                    continue
                descr = it.get("descr") or it.get("description") or ""
                if _is_consumable(name, descr, pid):
                    continue
                serial = it.get("sn") or it.get("serial_number") or ""
                eox = hw_eox.get(pid)
                status, days = _compute_status(eox)
                rows.append(self._make_hw_row(
                    hostname, mgmt_ip, name,
                    pid, eox.get("description", "") if eox else descr,
                    serial, eox, status, days,
                ))

        rows.sort(key=lambda r: (r[11] == "" or r[11] is None, r[11] if isinstance(r[11], int) else 999999))
        return rows

    def _build_software_rows(self, inventory_data, sw_eox):
        """Build rows for Software Lifecycle tab. One row per device."""
        from eox_client import EoxClient

        rows = []
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            device_type = device.get("device_type", "")
            if not device_type.startswith("cisco"):
                continue

            mgmt_ip = device.get("mgmt_ip", "")
            platform = device.get("platform", "")
            inv_data = device.get("collector_data", {}).get("device_inventory", {})
            parsed = inv_data.get("parsed", {})

            os_type = EoxClient.get_os_type(device_type)
            version_list = parsed.get("version", [])
            sw_ver = ""
            if version_list and isinstance(version_list, list):
                v = _normalize_keys(version_list[0])
                sw_ver = v.get("version") or v.get("os_version") or v.get("os") or ""
                if isinstance(sw_ver, list):
                    sw_ver = sw_ver[0] if sw_ver else ""
                if not platform:
                    platform = v.get("hardware") or v.get("platform") or ""
                    if isinstance(platform, list):
                        platform = platform[0] if platform else ""

            norm_ver = _normalize_sw_version(sw_ver, os_type) if sw_ver else sw_ver
            cache_key = f"{norm_ver}|{os_type}" if norm_ver and os_type else ""
            eox = sw_eox.get(cache_key) if cache_key else None
            status, days = _compute_status(eox)

            rows.append([
                hostname,
                mgmt_ip,
                device_type,
                platform,
                sw_ver,
                os_type or "",
                eox.get("end_of_sale", "") if eox else "",
                eox.get("end_of_sw_maintenance", "") if eox else "",
                eox.get("end_of_vulnerability_support", "") if eox else "",
                eox.get("last_date_of_support", "") if eox else "",
                days if days is not None else "",
                status,
                eox.get("bulletin_url", "") if eox else "",
            ])

        rows.sort(key=lambda r: (r[10] == "", r[10] if r[10] != "" else 999999))
        return rows

    def _make_hw_row(self, hostname, mgmt_ip, component, pid, description,
                     serial, eox, status, days):
        return [
            hostname,
            mgmt_ip,
            component,
            pid,
            description if eox and eox.get("description") else description,
            serial,
            eox.get("end_of_sale", "") if eox else "",
            eox.get("end_of_sw_maintenance", "") if eox else "",
            eox.get("end_of_vulnerability_support", "") if eox else "",
            eox.get("end_of_service_contract_renewal", "") if eox else "",
            eox.get("last_date_of_support", "") if eox else "",
            days if days is not None else "",
            status,
            eox.get("migration_pid", "") if eox else "",
            eox.get("migration_name", "") if eox else "",
            eox.get("bulletin_url", "") if eox else "",
        ]

    def _build_summary_sheet(self, wb, inventory_data, hw_rows, sw_rows, mod_rows, hw_eox, sw_eox):
        ws = wb.create_sheet(title="Summary")

        # Count Cisco vs non-Cisco devices
        total_devices = len(inventory_data.get("devices", {}))
        cisco_devices = sum(
            1 for d in inventory_data.get("devices", {}).values()
            if d.get("device_type", "").startswith("cisco")
        )
        non_cisco = total_devices - cisco_devices

        # Hardware stats
        unique_pids_checked = len(set(r[3] for r in hw_rows + mod_rows if r[3]))
        pids_with_data = len(set(r[3] for r in hw_rows + mod_rows if r[3] and r[12] != STATUS_NO_DATA))
        pids_no_data = unique_pids_checked - pids_with_data

        # Software stats
        unique_sw = len(set(r[4] for r in sw_rows if r[4]))
        sw_with_data = len(set(r[4] for r in sw_rows if r[4] and r[11] != STATUS_NO_DATA))
        sw_no_data = unique_sw - sw_with_data

        # Build summary rows
        row = 1
        title_font = Font(name="Calibri", size=14, bold=True)
        header_font = Font(name="Calibri", size=11, bold=True)

        ws.cell(row=row, column=1, value="EOX Lifecycle Summary").font = title_font
        row += 2

        ws.cell(row=row, column=1, value="OVERVIEW").font = header_font
        row += 1
        summary_data = [
            ("Total Devices", total_devices),
            ("Cisco Devices", cisco_devices),
            ("Non-Cisco Devices (skipped)", non_cisco),
            ("", ""),
            ("HARDWARE", ""),
            ("Unique PIDs Checked", unique_pids_checked),
            ("PIDs with EOX Data", pids_with_data),
            ("PIDs with No EOX Data", pids_no_data),
            ("", ""),
            ("SOFTWARE", ""),
            ("Unique Versions Checked", unique_sw),
            ("Versions with EOX Data", sw_with_data),
            ("Versions with No EOX Data", sw_no_data),
        ]
        for label, value in summary_data:
            if label:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="STATUS BREAKDOWN").font = header_font
        row += 1

        # Count by status
        status_headers = ["Status", "HW Chassis", "HW Modules", "Software"]
        for col, h in enumerate(status_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
        row += 1

        for status_label in [STATUS_EOS, STATUS_APPROACHING, STATUS_END_OF_SALE, STATUS_CURRENT, STATUS_NOT_IN_EOX, STATUS_NO_PID, STATUS_NO_DATA]:
            hw_count = sum(1 for r in hw_rows if r[12] == status_label)
            mod_count = sum(1 for r in mod_rows if r[12] == status_label)
            sw_count = sum(1 for r in sw_rows if r[11] == status_label)
            fill = STATUS_FILLS.get(status_label)

            ws.cell(row=row, column=1, value=status_label)
            ws.cell(row=row, column=2, value=hw_count)
            ws.cell(row=row, column=3, value=mod_count)
            ws.cell(row=row, column=4, value=sw_count)

            if fill:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = fill
            row += 1

        auto_width(ws)

    def _build_lifecycle_sheet(self, wb, title, rows, is_software=False):
        if is_software:
            headers = [
                "Device Name", "Management IP", "Device Type", "Platform",
                "Current Version", "OS Type", "End-of-Sale",
                "End-of-SW Maintenance", "End-of-Vulnerability Support",
                "Last Date of Support", "Days Remaining", "Status",
                "Bulletin URL",
            ]
            status_col = 12  # 1-indexed
        else:
            headers = [
                "Device Name", "Management IP", "Component", "PID",
                "Description", "Serial Number", "End-of-Sale",
                "End-of-SW Maintenance", "End-of-Vulnerability Support",
                "End-of-Service Contract Renewal", "Last Date of Support",
                "Days Remaining", "Status", "Replacement PID",
                "Replacement Name", "Bulletin URL",
            ]
            status_col = 13

        ws = wb.create_sheet(title=title)
        style_header_row(ws, headers)
        write_data_rows(ws, rows)

        # Apply status fills
        for row_idx in range(2, len(rows) + 2):
            status_cell = ws.cell(row=row_idx, column=status_col)
            status_val = status_cell.value
            fill = STATUS_FILLS.get(status_val)
            if fill:
                status_cell.fill = fill

            # Color days remaining
            days_col = status_col - 1
            days_cell = ws.cell(row=row_idx, column=days_col)
            if isinstance(days_cell.value, (int, float)):
                if days_cell.value < 0:
                    days_cell.fill = STATUS_FILLS[STATUS_EOS]
                elif days_cell.value <= APPROACHING_DAYS:
                    days_cell.fill = STATUS_FILLS[STATUS_APPROACHING]
                else:
                    days_cell.fill = STATUS_FILLS[STATUS_CURRENT]

        auto_width(ws)
        freeze_header(ws)
        add_table(ws, headers, len(rows))
