"""
MAC Table report -- MAC-to-port mappings per device.
"""
import io
from collections import OrderedDict

from openpyxl import Workbook

from .base import BaseReport
from .xlsx_utils import create_sheet


class MacTableReport(BaseReport):
    name = "mac_table"
    label = "MAC Table"
    description = "MAC address to port mappings per device"
    category = "Layer 2 Analysis"
    required_collectors = ["mac_table"]
    supported_formats = ["xlsx", "json", "csv", "xml"]

    def generate_tabular_data(self, inventory_data):
        headers = ["Device", "MAC Address", "VLAN", "Type", "Port"]
        rows = []
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            mac_data = device.get("collector_data", {}).get("mac_table", {})
            entries = mac_data.get("parsed", {}).get("entries", [])
            for entry in entries:
                rows.append([
                    hostname,
                    entry.get("mac", entry.get("destination_address", "")),
                    entry.get("vlan", entry.get("vlan_id", "")),
                    entry.get("type", ""),
                    entry.get("ports", entry.get("destination_port", "")),
                ])
        return OrderedDict([("MAC Table", (headers, rows))])

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        if fmt != "xlsx":
            return self._generate_for_format(inventory_data, fmt)

        wb = Workbook()
        wb.remove(wb.active)

        rows = []
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            mac_data = device.get("collector_data", {}).get("mac_table", {})
            entries = mac_data.get("parsed", {}).get("entries", [])
            for entry in entries:
                rows.append([
                    hostname,
                    entry.get("mac", entry.get("destination_address", "")),
                    entry.get("vlan", entry.get("vlan_id", "")),
                    entry.get("type", ""),
                    entry.get("ports", entry.get("destination_port", "")),
                ])

        create_sheet(wb, "MAC Table",
                     ["Device", "MAC Address", "VLAN", "Type", "Port"],
                     rows)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
