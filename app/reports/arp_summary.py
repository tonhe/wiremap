"""
ARP Summary report -- ARP tables by device with summary and detail tabs.
"""
import io
from collections import OrderedDict, defaultdict
from openpyxl import Workbook

from .base import BaseReport
from .xlsx_utils import create_sheet


class ArpSummaryReport(BaseReport):
    name = "arp_summary"
    label = "ARP Summary"
    description = "ARP table entries by device and interface"
    category = "Layer 3 & Routing"
    required_collectors = ["arp"]
    supported_formats = ["xlsx", "json", "csv", "xml"]

    def generate_tabular_data(self, inventory_data):
        detail_rows = []
        summary_counts = defaultdict(lambda: defaultdict(int))

        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            arp_data = device.get("collector_data", {}).get("arp", {})
            entries = arp_data.get("parsed", {}).get("entries", [])
            for entry in entries:
                intf = entry.get("interface", "")
                detail_rows.append([
                    hostname,
                    intf,
                    entry.get("ip", ""),
                    entry.get("mac", ""),
                    entry.get("age", ""),
                ])
                summary_counts[hostname][intf] += 1

        summary_rows = []
        for hostname in sorted(summary_counts.keys()):
            for intf in sorted(summary_counts[hostname].keys()):
                summary_rows.append([hostname, intf, summary_counts[hostname][intf]])

        return OrderedDict([
            ("Summary", (["Device", "Interface", "Entry Count"], summary_rows)),
            ("Detail", (["Device", "Interface", "IP", "MAC", "Age"], detail_rows)),
        ])

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        if fmt != "xlsx":
            return self._generate_for_format(inventory_data, fmt)

        tabular = self.generate_tabular_data(inventory_data)

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, (headers, rows) in tabular.items():
            create_sheet(wb, sheet_name, headers, rows)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
