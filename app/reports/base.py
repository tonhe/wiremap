"""
Base class for all reports.
Each report generates output from discovery inventory data.
"""
from abc import ABC, abstractmethod
from collections import OrderedDict
import csv
import io
import json
import xml.etree.ElementTree as ET

from .xlsx_utils import create_sheet


class BaseReport(ABC):
    """Abstract base for reports. Subclass and set class attributes."""

    name: str = ""
    label: str = ""
    description: str = ""
    required_collectors: list[str] = []
    category: str = "General"
    supported_formats: list[str] = ["xlsx"]

    @abstractmethod
    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        """Generate report from inventory data.

        Args:
            inventory_data: Full discovery inventory dict.
            fmt: Output format (must be in supported_formats).

        Returns:
            Report content as bytes.
        """

    def can_generate(self, inventory_data: dict) -> bool:
        """Check if inventory has the required collector data."""
        devices = inventory_data.get("devices", {})
        if not devices:
            return False
        for device in devices.values():
            collector_data = device.get("collector_data", {})
            if all(c in collector_data for c in self.required_collectors):
                return True
        return False

    def get_ui_options(self) -> list[dict]:
        """Return optional per-report config for the UI.

        Override in subclasses to expose report-specific toggles.
        Returns list of dicts: [{"name": "...", "label": "...", "type": "bool", "default": True}]
        """
        return []

    # ------------------------------------------------------------------
    # Multi-format support
    # ------------------------------------------------------------------

    def generate_tabular_data(self, inventory_data):
        """Return an OrderedDict of {sheet_name: (headers, rows)} or None.

        Subclasses override this to provide tabular data for multi-format
        export.  Returns None by default.
        """
        return None

    def _to_xlsx(self, tabular_data):
        from openpyxl import Workbook

        wb = Workbook()
        # Remove the default empty sheet created by openpyxl
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
        for sheet_name, (headers, rows) in tabular_data.items():
            create_sheet(wb, sheet_name, headers, rows)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _to_json(self, tabular_data):
        output = OrderedDict()
        for sheet_name, (headers, rows) in tabular_data.items():
            output[sheet_name] = [
                OrderedDict(zip(headers, row)) for row in rows
            ]
        return json.dumps(output, indent=2, default=str).encode("utf-8")

    def _to_csv(self, tabular_data):
        buf = io.StringIO()
        writer = csv.writer(buf)
        sheets = list(tabular_data.items())
        for idx, (sheet_name, (headers, rows)) in enumerate(sheets):
            if idx > 0:
                buf.write("\n")
            buf.write(f"--- Sheet: {sheet_name} ---\n")
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
        return buf.getvalue().encode("utf-8")

    def _to_xml(self, tabular_data):
        root = ET.Element("report")
        for sheet_name, (headers, rows) in tabular_data.items():
            sheet_el = ET.SubElement(root, "sheet", name=sheet_name)
            for row in rows:
                row_el = ET.SubElement(sheet_el, "row")
                for header, value in zip(headers, row):
                    field_el = ET.SubElement(row_el, "field", name=header)
                    field_el.text = str(value) if value is not None else ""
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def _generate_for_format(self, inventory_data, fmt):
        tabular_data = self.generate_tabular_data(inventory_data)
        if tabular_data is None:
            raise NotImplementedError(
                f"{self.__class__.__name__} does not support tabular export"
            )
        converters = {
            "xlsx": self._to_xlsx,
            "json": self._to_json,
            "csv": self._to_csv,
            "xml": self._to_xml,
        }
        converter = converters.get(fmt)
        if converter is None:
            raise ValueError(f"Unsupported format: {fmt}")
        return converter(tabular_data)
