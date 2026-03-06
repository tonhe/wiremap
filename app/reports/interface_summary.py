"""
Interface Summary report -- port status, descriptions, IP assignments.
"""
import io
from collections import OrderedDict

from openpyxl import Workbook

from .base import BaseReport
from .xlsx_utils import create_sheet


class InterfaceSummaryReport(BaseReport):
    name = "interface_summary"
    label = "Interface Summary"
    description = "Interface status, descriptions, and IP assignments"
    category = "Discovery & Topology"
    required_collectors = ["interfaces"]
    supported_formats = ["xlsx", "json", "csv", "xml"]

    def generate_tabular_data(self, inventory_data):
        status_headers = ["Device", "Port", "Description", "Status", "VLAN",
                          "Duplex", "Speed", "Type"]
        ip_headers = ["Device", "Interface", "IP Address", "Status", "Protocol"]

        status_rows = []
        ip_rows = []

        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            intf_data = device.get("collector_data", {}).get("interfaces", {})
            parsed = intf_data.get("parsed", {})

            # Interface status
            for entry in parsed.get("interfaces_status", []):
                status_rows.append([
                    hostname,
                    entry.get("port", entry.get("interface", "")),
                    entry.get("name", entry.get("description", "")),
                    entry.get("status", ""),
                    entry.get("vlan", ""),
                    entry.get("duplex", ""),
                    entry.get("speed", ""),
                    entry.get("type", ""),
                ])

            # IP interfaces
            for entry in parsed.get("ip_interfaces", []):
                ip_rows.append([
                    hostname,
                    entry.get("intf", entry.get("interface", "")),
                    entry.get("ipaddr", entry.get("ip_address", "")),
                    entry.get("status", ""),
                    entry.get("proto", entry.get("protocol", "")),
                ])

        tabular = OrderedDict()
        tabular["Interface Status"] = (status_headers, status_rows)
        tabular["IP Interfaces"] = (ip_headers, ip_rows)
        return tabular

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        if fmt == "xlsx":
            wb = Workbook()
            wb.remove(wb.active)

            tabular = self.generate_tabular_data(inventory_data)
            for sheet_name, (headers, rows) in tabular.items():
                create_sheet(wb, sheet_name, headers, rows)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        return self._generate_for_format(inventory_data, fmt)
