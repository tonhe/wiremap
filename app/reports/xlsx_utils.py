"""
Shared XLSX styling utilities for all report modules.
Provides consistent header styling, auto-width columns, frozen panes,
and color fills using openpyxl.
"""
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Regex to strip characters illegal in XML 1.0 (which openpyxl/XLSX requires)
_ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


# Color palette
HEADER_BG = "1F4E79"       # Dark blue
HEADER_FG = "FFFFFF"       # White text
ALT_ROW_BG = "F2F7FB"      # Light blue for alternating rows
BORDER_COLOR = "D9E2EC"    # Light gray border

# Fonts
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
DATA_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)

# Fills
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")

# Alignment
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="center")

# Border
THIN_BORDER = Border(
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def style_header_row(ws, headers: list[str], row: int = 1):
    """Write headers to a row and apply header styling."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_data_rows(ws, rows: list[list], start_row: int = 2):
    """Write data rows with alternating row colors and borders."""
    for row_idx, row_data in enumerate(rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            if isinstance(value, str):
                value = _ILLEGAL_XML_RE.sub('', value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if (row_idx - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def auto_width(ws, min_width: int = 8, max_width: int = 50):
    """Auto-size columns based on content. Call after writing all data."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def freeze_header(ws, row: int = 2):
    """Freeze panes so headers stay visible when scrolling."""
    ws.freeze_panes = f"A{row}"


def add_table(ws, headers: list[str], num_rows: int) -> None:
    """Wrap the data range in an Excel Table for sort/filter functionality."""
    if num_rows == 0:
        return
    last_col = get_column_letter(len(headers))
    last_row = num_rows + 1  # +1 for header row
    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', ws.title)[:64] or "Table1"
    table = Table(displayName=safe_name, ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def create_sheet(wb: Workbook, title: str, headers: list[str],
                 rows: list[list]) -> None:
    """Create a fully styled worksheet with headers, data, auto-width, and freeze."""
    ws = wb.create_sheet(title=title)
    style_header_row(ws, headers)
    write_data_rows(ws, rows)
    auto_width(ws)
    freeze_header(ws)
    add_table(ws, headers, len(rows))
