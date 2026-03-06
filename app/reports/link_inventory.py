"""
Link Inventory report -- all neighbor links from CDP/LLDP and L3.
"""
import io
from collections import OrderedDict

from openpyxl import Workbook

from .base import BaseReport
from .xlsx_utils import create_sheet


class LinkInventoryReport(BaseReport):
    name = "link_inventory"
    label = "Link Inventory"
    description = "All discovered neighbor links (CDP/LLDP and L3 routing)"
    category = "Discovery & Topology"
    required_collectors = ["cdp_lldp"]
    supported_formats = ["xlsx", "json", "csv", "xml"]

    def generate_tabular_data(self, inventory_data):
        # Collect all links
        all_links = []
        seen = set()
        for hostname, device in sorted(inventory_data.get("devices", {}).items()):
            cdp_data = device.get("collector_data", {}).get("cdp_lldp", {})
            neighbors = cdp_data.get("parsed", {}).get("neighbors", [])
            for n in neighbors:
                remote = n.get("remote_device", "Unknown")
                local_intf = n.get("local_intf", "?")
                remote_intf = n.get("remote_intf", "?")
                # Deduplicate bidirectional links
                key = tuple(sorted([
                    f"{hostname}:{local_intf}",
                    f"{remote}:{remote_intf}",
                ]))
                if key in seen:
                    continue
                seen.add(key)
                protocols = ", ".join(n.get("protocols", []))
                all_links.append([
                    hostname, local_intf, remote, remote_intf,
                    n.get("remote_ip", ""), protocols,
                ])

        headers = [
            "Local Device", "Local Interface", "Remote Device",
            "Remote Interface", "Remote IP", "Protocols",
        ]

        # By Device tab -- grouped
        by_device = {}
        for link in all_links:
            by_device.setdefault(link[0], []).append(link)
        grouped_rows = []
        for device_name in sorted(by_device.keys()):
            for link in by_device[device_name]:
                grouped_rows.append(link)

        tabular = OrderedDict()
        tabular["Links"] = (headers, all_links)
        tabular["By Device"] = (headers, grouped_rows)
        return tabular

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        if fmt == "xlsx":
            tabular = self.generate_tabular_data(inventory_data)
            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name, (headers, rows) in tabular.items():
                create_sheet(wb, sheet_name, headers, rows)
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        return self._generate_for_format(inventory_data, fmt)
