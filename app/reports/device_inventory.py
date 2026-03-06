"""
Device Inventory report -- hostname, IP, platform, version, serial.
Three sheets: Devices, Stack Members, Modules / Line Cards.
"""
import io
import re
from openpyxl import Workbook

from .base import BaseReport
from .xlsx_utils import create_sheet


def _normalize_keys(d: dict) -> dict:
    """Lowercase all dict keys for consistent field access."""
    return {k.lower(): v for k, v in d.items()}


# Common PID -> friendly platform description.
# Falls back to raw PID for unknown models.
_PLATFORM_DESCRIPTIONS = {
    # Cisco Catalyst 9000
    "C9200-24P": "Cisco Catalyst 9200 24-Port PoE",
    "C9200-48P": "Cisco Catalyst 9200 48-Port PoE",
    "C9200L-24P": "Cisco Catalyst 9200L 24-Port PoE",
    "C9200L-48P": "Cisco Catalyst 9200L 48-Port PoE",
    "C9300-24T": "Cisco Catalyst 9300 24-Port",
    "C9300-24P": "Cisco Catalyst 9300 24-Port PoE+",
    "C9300-48T": "Cisco Catalyst 9300 48-Port",
    "C9300-48P": "Cisco Catalyst 9300 48-Port PoE+",
    "C9300-48U": "Cisco Catalyst 9300 48-Port UPOE",
    "C9300L-24T": "Cisco Catalyst 9300L 24-Port",
    "C9300L-48T": "Cisco Catalyst 9300L 48-Port",
    "C9300L-48P": "Cisco Catalyst 9300L 48-Port PoE+",
    "C9300X-48HX": "Cisco Catalyst 9300X 48-Port mGig",
    "C9400": "Cisco Catalyst 9400 Series",
    "C9500-40X": "Cisco Catalyst 9500 40-Port 10G",
    "C9500-48Y4C": "Cisco Catalyst 9500 48-Port 25G",
    "C9600": "Cisco Catalyst 9600 Series",
    # Cisco Catalyst Classic
    "WS-C2960X-24": "Cisco Catalyst 2960-X 24-Port",
    "WS-C2960X-48": "Cisco Catalyst 2960-X 48-Port",
    "WS-C2960X-48FPD-L": "Cisco Catalyst 2960-X 48-Port PoE",
    "WS-C2960X-48FPS-L": "Cisco Catalyst 2960-X 48-Port PoE",
    "WS-C2960X-48LPD-L": "Cisco Catalyst 2960-X 48-Port PoE",
    "WS-C2960X-48LPS-L": "Cisco Catalyst 2960-X 48-Port PoE",
    "WS-C2960X-48TD-L": "Cisco Catalyst 2960-X 48-Port",
    "WS-C2960X-48TS-L": "Cisco Catalyst 2960-X 48-Port",
    "WS-C2960X-24PD-L": "Cisco Catalyst 2960-X 24-Port PoE",
    "WS-C2960X-24PS-L": "Cisco Catalyst 2960-X 24-Port PoE",
    "WS-C2960X-24TD-L": "Cisco Catalyst 2960-X 24-Port",
    "WS-C2960X-24TS-L": "Cisco Catalyst 2960-X 24-Port",
    "WS-C2960S-48": "Cisco Catalyst 2960-S 48-Port",
    "WS-C2960S-24": "Cisco Catalyst 2960-S 24-Port",
    "WS-C2960X-48FPD": "Cisco Catalyst 2960-X 48-Port PoE",
    "WS-C2960S-48FPD-L": "Cisco Catalyst 2960-S 48-Port PoE",
    "WS-C2960S-48LPD-L": "Cisco Catalyst 2960-S 48-Port PoE",
    "WS-C3560X-24": "Cisco Catalyst 3560-X 24-Port",
    "WS-C3560X-48": "Cisco Catalyst 3560-X 48-Port",
    "WS-C3750X-24": "Cisco Catalyst 3750-X 24-Port",
    "WS-C3750X-48": "Cisco Catalyst 3750-X 48-Port",
    "WS-C3750-48P": "Cisco Catalyst 3750 48-Port PoE",
    "WS-C3750-48P-S": "Cisco Catalyst 3750 48-Port PoE",
    "WS-C3850-24T": "Cisco Catalyst 3850 24-Port",
    "WS-C3850-48T": "Cisco Catalyst 3850 48-Port",
    "WS-C3850-48P": "Cisco Catalyst 3850 48-Port PoE+",
    "WS-C3850-24P": "Cisco Catalyst 3850 24-Port PoE+",
    "WS-C4500X-16": "Cisco Catalyst 4500-X 16-Port 10G",
    "WS-C4500X-32": "Cisco Catalyst 4500-X 32-Port 10G",
    # Cisco 2960 non-X
    "WS-C2960-48": "Cisco Catalyst 2960 48-Port",
    "WS-C2960-24": "Cisco Catalyst 2960 24-Port",
    "WS-C2960S-48FPS-L": "Cisco Catalyst 2960-S 48-Port PoE",
    "WS-C2960S-24PS-L": "Cisco Catalyst 2960-S 24-Port PoE",
    # Cisco Nexus
    "N5K-C5548UP": "Cisco Nexus 5548 Unified Ports",
    "N5K-C5596UP": "Cisco Nexus 5596 Unified Ports",
    "N5K-C56128P": "Cisco Nexus 56128P",
    "N7K-C7009": "Cisco Nexus 7009",
    "N7K-C7010": "Cisco Nexus 7010",
    "N7K-C7018": "Cisco Nexus 7018",
    "N9K-C9332PQ": "Cisco Nexus 9332PQ",
    "N9K-C9372PX": "Cisco Nexus 9372PX",
    "N9K-C93180YC-EX": "Cisco Nexus 93180YC-EX",
    "N9K-C93108TC-EX": "Cisco Nexus 93108TC-EX",
    "N9K-C9336C-FX2": "Cisco Nexus 9336C-FX2",
    # Cisco Routers / Voice Gateways
    "ISR4331": "Cisco ISR 4331",
    "ISR4351": "Cisco ISR 4351",
    "ISR4431": "Cisco ISR 4431",
    "ISR4451": "Cisco ISR 4451",
    "C1111-8P": "Cisco ISR 1100 8-Port",
    "CISCO2911/K9": "Cisco 2911",
    "Cisco VG310": "Cisco Voice Gateway 310",
    "Cisco VG320": "Cisco Voice Gateway 320",
    "Cisco VG224": "Cisco Voice Gateway 224",
    # Cisco ASR
    "ASR1001-X": "Cisco ASR 1001-X",
    "ASR1002-X": "Cisco ASR 1002-X",
    "ASR1006-X": "Cisco ASR 1006-X",
    # Cisco WAN / XR
    "WS-C2960XR-48FPD-I": "Cisco Catalyst 2960-XR 48-Port PoE",
    "NCS-5501": "Cisco NCS 5501",
    # Arista
    "DCS-7050TX-48": "Arista 7050TX 48-Port",
    "DCS-7050SX-64": "Arista 7050SX 64-Port",
    "DCS-7280SR-48C6": "Arista 7280R 48-Port",
    "DCS-7150S-24": "Arista 7150S 24-Port",
    "DCS-7050CX3-32S": "Arista 7050CX3 32-Port",
}

# Components to skip when falling back to show inventory
_SKIP_INVENTORY_TYPES = {
    "fan", "power supply", "psu", "sfp", "transceiver",
    "sensor", "clk", "usb",
}

# Patterns for models that are chassis/switch PIDs, not real modules.
_CHASSIS_MODEL_RE = re.compile(
    r"^("
    r"WS-C\d|"            # Catalyst 2960/3560/3750/3850/4500
    r"C9[2-6]\d{2}[LX]?-\d|"  # Catalyst 9200/9300/9400/9500/9600
    r"CISCO\d|"           # CISCO871, CISCO881, etc.
    r"ISR\d.*/K9|"        # ISR4331/K9, ISR4451-X/K9
    r"VG[23]\d{2}"        # VG224, VG310, VG320 chassis
    r")", re.IGNORECASE
)

# Patterns for module types/models to exclude.
_SKIP_MODULE_RE = re.compile(
    r"("
    r"stack|"             # stacking modules
    r"PVDM|"             # DSP modules
    r"GBIC|"             # GBICs
    r"built-in|"         # ISR built-in controllers/processors
    r"onboard.*voice|"   # onboard FXS voice interfaces
    r"ISR\d{4}\S*-\d+x\d+GE"  # ISR4331-3x1GE, ISR4451-X-4x1GE built-in ports
    r")", re.IGNORECASE
)


def _is_real_module(model, mod_type):
    """Return True if this module entry is a real line card / network module."""
    if _CHASSIS_MODEL_RE.match(model):
        return False
    if _SKIP_MODULE_RE.search(model) or _SKIP_MODULE_RE.search(mod_type):
        return False
    return True


def _get_platform_description(pid: str) -> str:
    """Look up friendly platform name from PID. Falls back to empty string."""
    if not pid:
        return ""
    if pid in _PLATFORM_DESCRIPTIONS:
        return _PLATFORM_DESCRIPTIONS[pid]
    # Try prefix match (strip trailing -S, -L suffixes)
    base = pid.rstrip("-").rsplit("-", 1)[0] if "-" in pid else pid
    if base in _PLATFORM_DESCRIPTIONS:
        return _PLATFORM_DESCRIPTIONS[base]
    # Try without /K9 suffix
    base = pid.split("/")[0]
    if base in _PLATFORM_DESCRIPTIONS:
        return _PLATFORM_DESCRIPTIONS[base]
    return ""


class DeviceInventoryReport(BaseReport):
    name = "device_inventory"
    label = "Device Inventory"
    description = "Device details: hostname, IP, platform, software version, serial numbers"
    category = "Discovery & Topology"
    required_collectors = ["device_inventory"]
    supported_formats = ["xlsx"]

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        self._build_devices_sheet(wb, inventory_data)
        self._build_stack_sheet(wb, inventory_data)
        self._build_modules_sheet(wb, inventory_data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _build_devices_sheet(self, wb, inventory_data):
        headers = [
            "Hostname", "Mgmt IP", "Device Type", "Category",
            "Platform", "Platform Description", "Software Version",
            "Serial Number",
        ]
        rows = []
        for hostname, device in sorted(
            inventory_data.get("devices", {}).items()
        ):
            inv_data = (device.get("collector_data", {})
                        .get("device_inventory", {}))
            parsed = inv_data.get("parsed", {})

            version_list = parsed.get("version", [])
            sw_version = ""
            serial = ""
            platform = device.get("platform", "")

            if (version_list and isinstance(version_list, list)
                    and len(version_list) > 0):
                v = _normalize_keys(version_list[0])
                sw_version = (v.get("version") or v.get("os_version")
                              or v.get("os") or "")
                serial = (v.get("serial") or v.get("serial_number")
                          or v.get("serialnum") or "")
                if not platform:
                    platform = (v.get("hardware")
                                or v.get("platform") or "")

            if not serial:
                inv_list = parsed.get("inventory", [])
                if (inv_list and isinstance(inv_list, list)
                        and len(inv_list) > 0):
                    vi = _normalize_keys(inv_list[0])
                    serial = (vi.get("sn")
                              or vi.get("serial_number") or "")

            if isinstance(platform, list):
                platform = ", ".join(str(x) for x in platform)
            if isinstance(sw_version, list):
                sw_version = ", ".join(str(x) for x in sw_version)
            if isinstance(serial, list):
                serial = ", ".join(str(x) for x in serial)

            rows.append([
                hostname,
                device.get("mgmt_ip", ""),
                device.get("device_type", ""),
                device.get("device_category", ""),
                platform,
                _get_platform_description(platform),
                sw_version,
                serial,
            ])

        create_sheet(wb, "Devices", headers, rows)

    def _build_stack_sheet(self, wb, inventory_data):
        headers = [
            "Hostname", "Mgmt IP", "Member #", "Role",
            "Model", "Serial", "MAC", "Priority", "State",
        ]
        rows = []
        for hostname, device in sorted(
            inventory_data.get("devices", {}).items()
        ):
            inv_data = (device.get("collector_data", {})
                        .get("device_inventory", {}))
            parsed = inv_data.get("parsed", {})
            stack = parsed.get("stack_members", [])
            if not stack:
                continue

            mgmt_ip = device.get("mgmt_ip", "")
            for member in stack:
                m = _normalize_keys(member)
                rows.append([
                    hostname,
                    mgmt_ip,
                    m.get("switch") or m.get("member") or "",
                    m.get("role") or "",
                    m.get("model") or m.get("hw_ver") or "",
                    m.get("serial") or m.get("serialnum") or "",
                    m.get("mac_address") or m.get("mac") or "",
                    m.get("priority") or "",
                    m.get("state") or m.get("status") or "",
                ])

        create_sheet(wb, "Stack Members", headers, rows)

    def _build_modules_sheet(self, wb, inventory_data):
        headers = [
            "Hostname", "Mgmt IP", "Slot", "Type",
            "Model", "Serial", "Status", "Ports",
        ]
        rows = []
        for hostname, device in sorted(
            inventory_data.get("devices", {}).items()
        ):
            inv_data = (device.get("collector_data", {})
                        .get("device_inventory", {}))
            parsed = inv_data.get("parsed", {})
            modules = parsed.get("modules", [])
            mgmt_ip = device.get("mgmt_ip", "")

            if modules:
                for mod in modules:
                    m = _normalize_keys(mod)
                    model = m.get("model") or ""
                    mod_type = m.get("type") or m.get("module_type") or ""
                    if not _is_real_module(model, mod_type):
                        continue
                    rows.append([
                        hostname,
                        mgmt_ip,
                        m.get("module") or m.get("slot") or "",
                        mod_type,
                        model,
                        m.get("serial") or m.get("serialnum") or "",
                        m.get("status") or "",
                        m.get("ports") or "",
                    ])
            else:
                inv_list = parsed.get("inventory", [])
                for item in inv_list:
                    it = _normalize_keys(item)
                    descr = (it.get("descr")
                             or it.get("description") or "").lower()
                    name = (it.get("name") or "").lower()
                    if any(skip in descr or skip in name
                           for skip in _SKIP_INVENTORY_TYPES):
                        continue
                    if "chassis" in name:
                        continue
                    rows.append([
                        hostname,
                        mgmt_ip,
                        it.get("name") or "",
                        it.get("descr") or it.get("description") or "",
                        it.get("pid") or "",
                        it.get("sn") or it.get("serial_number") or "",
                        "",
                        "",
                    ])

        create_sheet(wb, "Modules", headers, rows)
