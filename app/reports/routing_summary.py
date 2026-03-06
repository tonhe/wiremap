"""
L3 Routing & IP Workbook -- comprehensive Layer 3 report.

Tabs:
  1. Summary -- high-level metrics and finding counts
  2. Protocol Neighbors -- OSPF, EIGRP, BGP, IS-IS neighbors
  3. Routes -- full routing table per device
  4. Routed Interfaces -- non-SVI routed interfaces
  5. Route Summary -- route counts by protocol per device
  6. OSPF Topology -- OSPF process, area, interface detail
  7. IP Address Audit -- all IPs with overlap detection
  8. ARP/MAC Map -- ARP entries correlated with MAC table
  9. VRF Summary -- VRF definitions and interface assignments
  10. Findings -- auto-detected L3 issues
"""
import io
import re
from collections import defaultdict
from ipaddress import IPv4Network, AddressValueError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .base import BaseReport
from .l2_discovery import (
    _extract_ip_interfaces,
    _get_interface_description,
    _get_cdp_neighbor,
    _derive_site,
    _has_collector_data,
)

# ---------------------------------------------------------------------------
# Styling (mirrors L2 discovery report palette)
# ---------------------------------------------------------------------------
HEADER_BG = "1F3864"
WHITE = "FFFFFF"
BLACK = "000000"
ALT_ROW_BG = "F2F2F2"
BORDER_COLOR = "D9E2EC"
FINDING_CRITICAL_BG = "FFC7CE"
FINDING_CRITICAL_FG = "9C0006"
FINDING_WARNING_BG = "FFF2CC"
FINDING_WARNING_FG = "7F6000"

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_DATA_FONT = Font(name="Calibri", size=11, color=BLACK)
_HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
_ALT_FILL = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
_WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="top")
_WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(bottom=Side(style="thin", color=BORDER_COLOR))

_CRITICAL_FONT = Font(name="Calibri", size=11, color=FINDING_CRITICAL_FG, bold=True)
_CRITICAL_FILL = PatternFill(start_color=FINDING_CRITICAL_BG, end_color=FINDING_CRITICAL_BG, fill_type="solid")
_WARNING_FONT = Font(name="Calibri", size=11, color=FINDING_WARNING_FG)
_WARNING_FILL = PatternFill(start_color=FINDING_WARNING_BG, end_color=FINDING_WARNING_BG, fill_type="solid")

# Regex to strip illegal XML characters
_ILLEGAL_XML_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _clean(value):
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub('', value)
    return value


def _write_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _write_row(ws, row_idx, values, data_start=2):
    fill = _ALT_FILL if (row_idx - data_start) % 2 == 1 else _WHITE_FILL
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=_clean(value))
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGN
        cell.border = _THIN_BORDER
        cell.fill = fill


def _auto_width(ws, min_width=8, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def _add_table(ws, table_name, num_cols, last_row):
    if last_row < 2:
        last_row = 2
    end_col = get_column_letter(num_cols)
    ref = f"A1:{end_col}{last_row}"
    style = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = style
    ws.add_table(table)


def _finalize(ws, table_name, num_cols, last_row):
    _add_table(ws, table_name, num_cols, last_row)
    _auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Analysis -- gather data and detect findings
# ---------------------------------------------------------------------------
def _analyze_l3(inventory_data):
    """Analyze L3 data across all devices. Returns structured analysis dict."""
    findings = []
    all_subnets = []  # (network_obj, hostname, interface, ip_str)
    ospf_neighbors_by_device = {}

    devices = inventory_data.get("devices", {})

    for hostname, device in sorted(devices.items()):
        cd = device.get("collector_data", {})

        # --- OSPF neighbor state check ---
        l3_parsed = cd.get("l3_routing", {}).get("parsed", {})
        for n in l3_parsed.get("neighbors", []):
            protocols = n.get("protocols", [])
            if "ospf" in protocols:
                # l3_routing stores parsed neighbors from 'show ip ospf neighbor'
                # The raw state comes through in the neighbor dict
                state = n.get("state", "")
                if state and "full" not in state.lower() and "dr" not in state.lower():
                    findings.append({
                        "hostname": hostname,
                        "severity": "Critical",
                        "title": "OSPF neighbor not FULL",
                        "description": (
                            f"Neighbor {n.get('remote_ip', '?')} state: {state}"
                        ),
                    })

        # --- EIGRP stub check ---
        rd_parsed = cd.get("routing_detail", {}).get("parsed", {})
        for en in rd_parsed.get("eigrp_neighbors", []):
            if en.get("stub"):
                # Stub is informational unless it's unexpected on a core router
                findings.append({
                    "hostname": hostname,
                    "severity": "Info",
                    "title": "EIGRP stub peer",
                    "description": (
                        f"Neighbor {en.get('neighbor', '?')} on {en.get('interface', '?')} "
                        f"is stub ({en.get('stub_flags', 'unknown flags')})"
                    ),
                })

        # --- Route summary: static routes alongside dynamic ---
        route_summary = rd_parsed.get("route_summary", [])
        has_dynamic = False
        has_static = False
        static_count = 0
        for rs in route_summary:
            source = str(rs.get("source", "")).lower()
            count = rs.get("count", 0)
            if not isinstance(count, int):
                try:
                    count = int(count)
                except (ValueError, TypeError):
                    count = 0
            if source in ("ospf", "eigrp", "bgp", "isis") or "ospf" in source or "eigrp" in source:
                if count > 0:
                    has_dynamic = True
            if source == "static" and count > 0:
                has_static = True
                static_count = count

        if has_dynamic and has_static:
            findings.append({
                "hostname": hostname,
                "severity": "Warning",
                "title": "Static routes alongside dynamic protocol",
                "description": (
                    f"{static_count} static route(s) present despite dynamic "
                    f"routing protocol(s) -- potential oversight"
                ),
            })

        # --- /32 host route leak check ---
        routes = l3_parsed.get("routes", [])
        host_routes = 0
        for r in routes:
            mask = str(r.get("mask", r.get("prefix_length", "")))
            if mask in ("32", "255.255.255.255", "/32"):
                host_routes += 1
        if host_routes > 20:
            findings.append({
                "hostname": hostname,
                "severity": "Warning",
                "title": "Large number of /32 host routes",
                "description": (
                    f"{host_routes} host routes detected -- "
                    f"possible redistribution leak"
                ),
            })

        # --- BGP peer state check ---
        for bp in rd_parsed.get("bgp_summary", []):
            state = bp.get("state", "")
            if state and state != "Established":
                findings.append({
                    "hostname": hostname,
                    "severity": "Critical",
                    "title": "BGP peer not Established",
                    "description": (
                        f"Neighbor {bp.get('neighbor', '?')} "
                        f"AS {bp.get('asn', '?')} state: {state}"
                    ),
                })

        # --- Collect IP subnets for overlap detection ---
        ip_interfaces = _extract_ip_interfaces(device)
        for intf in ip_interfaces:
            iface_name = intf["interface"]
            for ip_info in intf["ips"]:
                ip = ip_info["ip"]
                prefix = ip_info["prefix"]
                if ip and prefix:
                    try:
                        net = IPv4Network(f"{ip}/{prefix}", strict=False)
                        all_subnets.append((net, hostname, iface_name, f"{ip}/{prefix}"))
                    except (AddressValueError, ValueError):
                        pass

        # --- Collect OSPF neighbors for asymmetry check ---
        ospf_intfs = rd_parsed.get("ospf_interfaces", [])
        if ospf_intfs:
            ospf_neighbors_by_device[hostname] = ospf_intfs

    # --- Overlapping subnet detection ---
    # Compare all subnets; flag overlaps between different devices
    overlap_checked = set()
    for i, (net_a, host_a, intf_a, ip_a) in enumerate(all_subnets):
        for j, (net_b, host_b, intf_b, ip_b) in enumerate(all_subnets):
            if j <= i:
                continue
            if host_a == host_b:
                continue
            if net_a.overlaps(net_b):
                pair_key = tuple(sorted([(host_a, ip_a), (host_b, ip_b)]))
                if pair_key in overlap_checked:
                    continue
                overlap_checked.add(pair_key)
                # Same subnet on a shared link is fine; overlaps of different
                # size subnets are the real problem
                if net_a == net_b:
                    continue
                findings.append({
                    "hostname": host_a,
                    "severity": "Critical",
                    "title": "Overlapping IP subnet",
                    "description": (
                        f"{host_a} {intf_a} ({ip_a}) overlaps with "
                        f"{host_b} {intf_b} ({ip_b})"
                    ),
                })

    # --- Asymmetric OSPF cost detection ---
    # Check if interfaces in same area have wildly different costs
    area_costs = defaultdict(list)  # {area: [(hostname, intf, cost)]}
    for hostname, intfs in ospf_neighbors_by_device.items():
        for oi in intfs:
            area = oi.get("area", "")
            cost = oi.get("cost", "")
            try:
                cost_int = int(cost)
            except (ValueError, TypeError):
                continue
            area_costs[area].append((hostname, oi.get("interface", ""), cost_int))

    for area, entries in area_costs.items():
        if len(entries) < 2:
            continue
        costs = [e[2] for e in entries]
        min_cost, max_cost = min(costs), max(costs)
        if max_cost > 0 and min_cost > 0 and max_cost / min_cost > 100:
            findings.append({
                "hostname": entries[0][0],
                "severity": "Warning",
                "title": f"Asymmetric OSPF costs in area {area}",
                "description": (
                    f"Cost range {min_cost}-{max_cost} across "
                    f"{len(entries)} interfaces in area {area}"
                ),
            })

    return {"findings": findings}


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_summary_sheet(wb, analysis, inventory_data):
    ws = wb.create_sheet(title="Summary")
    headers = ["Metric", "Value"]
    _write_header_row(ws, headers)

    devices = inventory_data.get("devices", {})
    device_count = len(devices)

    # Count routed devices (those with l3_routing or routing_detail data)
    routed_count = 0
    total_routes = 0
    for device in devices.values():
        cd = device.get("collector_data", {})
        l3 = cd.get("l3_routing", {}).get("parsed", {})
        rd = cd.get("routing_detail", {}).get("parsed", {})
        if l3.get("neighbors") or l3.get("routes") or rd.get("route_summary"):
            routed_count += 1
        total_routes += len(l3.get("routes", []))

    # Finding counts
    critical = sum(1 for f in analysis["findings"] if f.get("severity") == "Critical")
    warning = sum(1 for f in analysis["findings"] if f.get("severity") == "Warning")
    info = sum(1 for f in analysis["findings"] if f.get("severity") == "Info")

    metrics = [
        ("Total Devices", device_count),
        ("Routed Devices", routed_count),
        ("Total Routes", total_routes),
        ("Critical Findings", critical),
        ("Warning Findings", warning),
        ("Info Findings", info),
    ]

    row = 2
    for metric, value in metrics:
        _write_row(ws, row, [metric, value])
        row += 1

    _finalize(ws, "L3Summary", len(headers), row - 1)


def _build_neighbors_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="Protocol Neighbors")
    headers = ["Device", "Site/Location", "Protocol", "Neighbor IP",
               "Neighbor Device", "State"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        l3_data = device.get("collector_data", {}).get("l3_routing", {})
        parsed = l3_data.get("parsed", {})
        site = _derive_site(hostname)

        for n in parsed.get("neighbors", []):
            _write_row(ws, row, [
                hostname,
                site,
                ", ".join(n.get("protocols", [])),
                n.get("remote_ip", ""),
                n.get("remote_device", ""),
                n.get("state", ""),
            ])
            row += 1

    _finalize(ws, "ProtocolNeighbors", len(headers), row - 1)


def _build_routes_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="Routes")
    headers = ["Device", "Site/Location", "Network", "Mask",
               "Next Hop", "Interface", "Protocol", "Metric"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        l3_data = device.get("collector_data", {}).get("l3_routing", {})
        parsed = l3_data.get("parsed", {})
        site = _derive_site(hostname)

        for r in parsed.get("routes", []):
            _write_row(ws, row, [
                hostname,
                site,
                r.get("network", r.get("destination", "")),
                r.get("mask", r.get("prefix_length", "")),
                r.get("nexthop_ip", r.get("next_hop", "")),
                r.get("nexthop_if", r.get("interface", "")),
                r.get("protocol", ""),
                r.get("metric", ""),
            ])
            row += 1

    _finalize(ws, "Routes", len(headers), row - 1)


def _build_routed_interfaces_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="Routed Interfaces")
    headers = ["Device", "Site/Location", "Interface", "IP / CIDR",
               "Description / Neighbor", "Notes"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        interfaces = _extract_ip_interfaces(device)
        site = _derive_site(hostname)

        for intf in sorted(interfaces, key=lambda x: x["interface"]):
            iface_name = intf["interface"]
            if re.match(r"^[Vv]lan\d+$", iface_name):
                continue
            for ip_info in intf["ips"]:
                cidr = (f"{ip_info['ip']}/{ip_info['prefix']}"
                        if ip_info["prefix"] else ip_info["ip"])
                desc = _get_interface_description(device, iface_name)
                if not desc:
                    desc = _get_cdp_neighbor(device, iface_name)
                iface_lower = iface_name.lower()
                if "loopback" in iface_lower:
                    note = "Loopback"
                elif "mgmt" in iface_lower or "management" in iface_lower:
                    note = "OOB management"
                elif "port-channel" in iface_lower or iface_lower[:2] == "po":
                    note = "Routed port-channel"
                elif "tunnel" in iface_lower:
                    note = "Tunnel"
                else:
                    note = "Routed interface"
                if ip_info["secondary"]:
                    note += " (secondary)"
                _write_row(ws, row, [hostname, site, iface_name, cidr, desc, note])
                row += 1

    _finalize(ws, "RoutedInterfaces", len(headers), row - 1)


def _build_route_summary_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="Route Summary")
    headers = ["Device", "Site/Location", "Connected", "Static",
               "OSPF", "EIGRP", "BGP", "Other", "Total"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        rd_parsed = device.get("collector_data", {}).get("routing_detail", {}).get("parsed", {})
        route_summary = rd_parsed.get("route_summary", [])
        if not route_summary:
            continue

        site = _derive_site(hostname)
        counts = {"connected": 0, "static": 0, "ospf": 0, "eigrp": 0, "bgp": 0, "other": 0}
        total = 0

        for rs in route_summary:
            source = str(rs.get("source", "")).lower()
            count = rs.get("count", 0)
            if not isinstance(count, int):
                try:
                    count = int(count)
                except (ValueError, TypeError):
                    count = 0

            if "connected" in source:
                counts["connected"] += count
            elif "static" in source:
                counts["static"] += count
            elif "ospf" in source:
                counts["ospf"] += count
            elif "eigrp" in source:
                counts["eigrp"] += count
            elif "bgp" in source:
                counts["bgp"] += count
            else:
                counts["other"] += count
            total += count

        _write_row(ws, row, [
            hostname, site,
            counts["connected"], counts["static"], counts["ospf"],
            counts["eigrp"], counts["bgp"], counts["other"], total,
        ])
        row += 1

    _finalize(ws, "RouteSummary", len(headers), row - 1)


def _build_ospf_topology_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="OSPF Topology")
    headers = ["Device", "Site/Location", "Process ID", "Router ID",
               "Area", "Interface", "Cost", "State", "Neighbors"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        rd_parsed = device.get("collector_data", {}).get("routing_detail", {}).get("parsed", {})
        ospf_procs = rd_parsed.get("ospf_processes", [])
        ospf_intfs = rd_parsed.get("ospf_interfaces", [])
        site = _derive_site(hostname)

        if not ospf_procs and not ospf_intfs:
            continue

        # Build process/area lookup
        proc_id = ""
        router_id = ""
        proc_areas = []
        if ospf_procs:
            p = ospf_procs[0]
            proc_id = p.get("process_id", "")
            router_id = p.get("router_id", "")
            proc_areas = p.get("areas", [])

        if ospf_intfs:
            for oi in ospf_intfs:
                _write_row(ws, row, [
                    hostname, site, proc_id, router_id,
                    oi.get("area", ""),
                    oi.get("interface", ""),
                    oi.get("cost", ""),
                    oi.get("state", ""),
                    oi.get("neighbors", oi.get("nbrs_full", "")),
                ])
                row += 1
        elif ospf_procs:
            # No interface detail, just list process + areas
            for area in proc_areas:
                _write_row(ws, row, [
                    hostname, site, proc_id, router_id,
                    area, "", "", "", "",
                ])
                row += 1

    _finalize(ws, "OSPFTopology", len(headers), row - 1)


def _build_ip_audit_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="IP Address Audit")
    headers = ["Device", "Site/Location", "Interface", "IP / CIDR",
               "Subnet", "VLAN", "Overlap"]
    _write_header_row(ws, headers)

    # Collect all subnets for overlap detection
    all_entries = []
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        ip_interfaces = _extract_ip_interfaces(device)
        site = _derive_site(hostname)

        for intf in ip_interfaces:
            iface_name = intf["interface"]
            # Try to determine VLAN from SVI name
            vlan = ""
            m = re.match(r"^[Vv]lan(\d+)$", iface_name)
            if m:
                vlan = m.group(1)

            for ip_info in intf["ips"]:
                ip = ip_info["ip"]
                prefix = ip_info["prefix"]
                cidr = f"{ip}/{prefix}" if prefix else ip
                subnet = ""
                net_obj = None
                if ip and prefix:
                    try:
                        net_obj = IPv4Network(f"{ip}/{prefix}", strict=False)
                        subnet = str(net_obj)
                    except (AddressValueError, ValueError):
                        pass

                all_entries.append({
                    "hostname": hostname,
                    "site": site,
                    "interface": iface_name,
                    "cidr": cidr,
                    "subnet": subnet,
                    "vlan": vlan,
                    "net_obj": net_obj,
                })

    # Detect overlaps
    overlap_map = {}  # index -> overlap description
    for i, a in enumerate(all_entries):
        if not a["net_obj"]:
            continue
        for j, b in enumerate(all_entries):
            if j <= i or not b["net_obj"]:
                continue
            if a["hostname"] == b["hostname"]:
                continue
            if a["net_obj"] == b["net_obj"]:
                continue  # same subnet on shared link is fine
            if a["net_obj"].overlaps(b["net_obj"]):
                desc = f"Overlaps {b['hostname']} {b['interface']}"
                overlap_map.setdefault(i, []).append(desc)
                desc_b = f"Overlaps {a['hostname']} {a['interface']}"
                overlap_map.setdefault(j, []).append(desc_b)

    row = 2
    for idx, entry in enumerate(all_entries):
        overlap = "; ".join(overlap_map.get(idx, []))
        _write_row(ws, row, [
            entry["hostname"], entry["site"], entry["interface"],
            entry["cidr"], entry["subnet"], entry["vlan"],
            overlap or "",
        ])
        # Highlight overlap rows
        if overlap:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _CRITICAL_FONT
                cell.fill = _CRITICAL_FILL
        row += 1

    _finalize(ws, "IPAddressAudit", len(headers), row - 1)


def _build_arp_mac_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="ARP MAC Map")
    headers = ["Device", "Site/Location", "IP Address", "MAC Address",
               "Interface", "VLAN", "Vendor OUI"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        cd = device.get("collector_data", {})
        arp_parsed = cd.get("arp", {}).get("parsed", {})
        mac_parsed = cd.get("mac_table", {}).get("parsed", {})
        site = _derive_site(hostname)

        # Build MAC -> VLAN lookup from mac_table
        mac_vlan = {}
        for me in mac_parsed.get("entries", []):
            mac = me.get("mac", me.get("destination_address", "")).lower()
            vlan = me.get("vlan", me.get("vlan_id", ""))
            if mac:
                mac_vlan[mac] = str(vlan)

        for entry in arp_parsed.get("entries", []):
            ip = entry.get("ip", entry.get("address", ""))
            mac = entry.get("mac", entry.get("hardware", ""))
            intf = entry.get("interface", entry.get("port", ""))
            vlan = mac_vlan.get(mac.lower(), "") if mac else ""

            # Extract OUI (first 3 octets)
            oui = ""
            if mac:
                clean_mac = mac.replace(":", "").replace(".", "").replace("-", "")
                if len(clean_mac) >= 6:
                    oui = clean_mac[:6].upper()

            _write_row(ws, row, [hostname, site, ip, mac, intf, vlan, oui])
            row += 1

    _finalize(ws, "ARPMACMap", len(headers), row - 1)


def _build_vrf_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="VRF Summary")
    headers = ["Device", "Site/Location", "VRF Name", "RD",
               "Interfaces", "Interface Count"]
    _write_header_row(ws, headers)

    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        vrf_parsed = device.get("collector_data", {}).get("vrf", {}).get("parsed", {})
        vrfs = vrf_parsed.get("vrfs", [])
        if not vrfs:
            continue

        site = _derive_site(hostname)
        for v in vrfs:
            intfs = v.get("interfaces", [])
            _write_row(ws, row, [
                hostname, site,
                v.get("name", ""),
                v.get("rd", ""),
                ", ".join(intfs),
                len(intfs),
            ])
            row += 1

    _finalize(ws, "VRFSummary", len(headers), row - 1)


def _build_findings_sheet(wb, analysis):
    ws = wb.create_sheet(title="Findings")
    headers = ["Device", "Site/Location", "Severity", "Finding", "Details"]
    _write_header_row(ws, headers)

    row = 2
    for finding in analysis["findings"]:
        hostname = finding.get("hostname", "")
        severity = finding.get("severity", "Info")
        _write_row(ws, row, [
            hostname,
            _derive_site(hostname),
            severity,
            finding["title"],
            finding["description"],
        ])
        # Color-code by severity
        if severity == "Critical":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _CRITICAL_FONT
                cell.fill = _CRITICAL_FILL
        elif severity == "Warning":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _WARNING_FONT
                cell.fill = _WARNING_FILL
        row += 1

    if not analysis["findings"]:
        ws.cell(row=2, column=1, value="No findings detected.")
        row = 3

    _finalize(ws, "L3Findings", len(headers), row - 1)


# ---------------------------------------------------------------------------
# Report class
# ---------------------------------------------------------------------------
class RoutingSummaryReport(BaseReport):
    name = "routing_summary"
    label = "L3 Routing & IP Report"
    description = "Protocol neighbors, route tables, OSPF topology, IP audit, ARP/MAC map, VRF summary, and findings"
    category = "Layer 3 & Routing"
    required_collectors = ["l3_routing"]
    supported_formats = ["xlsx"]

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        analysis = _analyze_l3(inventory_data)

        wb = Workbook()
        wb.remove(wb.active)

        # Tab 1: Summary
        _build_summary_sheet(wb, analysis, inventory_data)

        # Tab 2: Protocol Neighbors
        _build_neighbors_sheet(wb, inventory_data)

        # Tab 3: Routes
        _build_routes_sheet(wb, inventory_data)

        # Tab 4: Routed Interfaces
        _build_routed_interfaces_sheet(wb, inventory_data)

        # Tab 5: Route Summary (if routing_detail data exists)
        if _has_collector_data(inventory_data, "routing_detail"):
            _build_route_summary_sheet(wb, inventory_data)

        # Tab 6: OSPF Topology (if routing_detail data exists)
        if _has_collector_data(inventory_data, "routing_detail"):
            _build_ospf_topology_sheet(wb, inventory_data)

        # Tab 7: IP Address Audit (if interfaces data exists)
        if _has_collector_data(inventory_data, "interfaces"):
            _build_ip_audit_sheet(wb, inventory_data)

        # Tab 8: ARP/MAC Map (if arp data exists)
        if _has_collector_data(inventory_data, "arp"):
            _build_arp_mac_sheet(wb, inventory_data)

        # Tab 9: VRF Summary (if vrf data exists)
        if _has_collector_data(inventory_data, "vrf"):
            _build_vrf_sheet(wb, inventory_data)

        # Tab 10: Findings
        _build_findings_sheet(wb, analysis)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
