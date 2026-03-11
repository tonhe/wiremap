"""
Internet Edge report -- BGP peering, NAT, ACL audit, FHRP, edge interfaces.
"""
import io
from collections import OrderedDict

from openpyxl import Workbook

from .base import BaseReport
from .l2_discovery import _derive_site, _has_collector_data
from .routing_summary import (
    HEADER_BG, _write_header_row, _write_row, _finalize,
    _CRITICAL_FONT, _CRITICAL_FILL, _WARNING_FONT, _WARNING_FILL,
)


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def _analyze_edge(inventory_data):
    findings = []
    devices = inventory_data.get("devices", {})

    for hostname, device in sorted(devices.items()):
        cd = device.get("collector_data", {})

        # --- BGP peer state ---
        rd_parsed = cd.get("routing_detail", {}).get("parsed", {})
        bgp_peers = rd_parsed.get("bgp_summary", [])

        for bp in bgp_peers:
            state = bp.get("state", "")
            if state and state != "Established":
                findings.append({
                    "hostname": hostname,
                    "severity": "Critical",
                    "title": "BGP peer not Established",
                    "description": (
                        f"Neighbor {bp.get('neighbor', '?')} "
                        f"AS {bp.get('asn', '?')} state: {state}"
                    ),
                })

        # --- Single-homed BGP ---
        if len(bgp_peers) == 1:
            findings.append({
                "hostname": hostname,
                "severity": "Warning",
                "title": "Single-homed BGP",
                "description": (
                    f"Only 1 BGP peer ({bgp_peers[0].get('neighbor', '?')} "
                    f"AS {bgp_peers[0].get('asn', '?')}) -- no redundancy"
                ),
            })

        # --- NAT pool utilization ---
        es_parsed = cd.get("edge_services", {}).get("parsed", {})
        nat_stats = es_parsed.get("nat_statistics", {})
        for pool in nat_stats.get("pools", []):
            util = pool.get("utilization_pct", 0)
            if isinstance(util, (int, float)) and util > 80:
                findings.append({
                    "hostname": hostname,
                    "severity": "Warning",
                    "title": "NAT pool high utilization",
                    "description": (
                        f"Pool '{pool.get('name', '?')}' at {util}% "
                        f"({pool.get('allocated', '?')}/{pool.get('total_addresses', '?')})"
                    ),
                })

        # --- ACL audit ---
        for acl in es_parsed.get("access_lists", []):
            entries = acl.get("entries", [])
            if not entries:
                continue
            # Dead ACL (all zero hits)
            all_zero = all(e.get("hit_count", 0) == 0 for e in entries)
            if all_zero:
                findings.append({
                    "hostname": hostname,
                    "severity": "Info",
                    "title": "Dead ACL (all zero hits)",
                    "description": f"ACL '{acl.get('name', '?')}' has {len(entries)} rules, all with zero hits",
                })
            # Permit any any
            for e in entries:
                if (e.get("action", "").lower() == "permit"
                        and e.get("protocol", "").lower() == "ip"
                        and e.get("source", "").lower() == "any"
                        and e.get("destination", "").lower() == "any"):
                    findings.append({
                        "hostname": hostname,
                        "severity": "Warning",
                        "title": "ACL permit ip any any",
                        "description": f"ACL '{acl.get('name', '?')}' contains permit ip any any",
                    })

        # --- FHRP unexpected state ---
        hsrp_parsed = cd.get("hsrp", {}).get("parsed", {})
        for e in hsrp_parsed.get("entries", []):
            state = e.get("state", "")
            if state and state.lower() not in ("active", "standby", "listen", ""):
                findings.append({
                    "hostname": hostname,
                    "severity": "Warning",
                    "title": "FHRP group in unexpected state",
                    "description": (
                        f"Interface {e.get('interface', '?')} group {e.get('group', '?')} "
                        f"VIP {e.get('virtual_ip', '?')} state: {state}"
                    ),
                })

    return {"findings": findings}


# ---------------------------------------------------------------------------
# XLSX sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb, analysis, inventory_data):
    ws = wb.create_sheet(title="Summary")
    headers = ["Metric", "Value"]
    _write_header_row(ws, headers)

    devices = inventory_data.get("devices", {})
    edge_count = 0
    bgp_peer_count = 0
    nat_active = 0

    for device in devices.values():
        cd = device.get("collector_data", {})
        rd = cd.get("routing_detail", {}).get("parsed", {})
        es = cd.get("edge_services", {}).get("parsed", {})
        bgp = rd.get("bgp_summary", [])
        if bgp or es.get("nat_statistics") or es.get("access_lists"):
            edge_count += 1
        bgp_peer_count += len(bgp)
        nat_stats = es.get("nat_statistics", {})
        nat_active += nat_stats.get("active_translations", 0)

    critical = sum(1 for f in analysis["findings"] if f.get("severity") == "Critical")
    warning = sum(1 for f in analysis["findings"] if f.get("severity") == "Warning")
    info = sum(1 for f in analysis["findings"] if f.get("severity") == "Info")

    rows = [
        ["Total Devices", len(devices)],
        ["Edge Devices", edge_count],
        ["BGP Peers (total)", bgp_peer_count],
        ["Active NAT Translations", nat_active],
        ["Critical Findings", critical],
        ["Warning Findings", warning],
        ["Info Findings", info],
    ]
    for i, row_data in enumerate(rows, 2):
        _write_row(ws, i, row_data)
    _finalize(ws, "EdgeSummary", len(headers), len(rows) + 1)


def _build_bgp_peers_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="BGP Peers")
    headers = ["Device", "Site/Location", "Neighbor", "ASN", "State",
               "Prefixes Rcvd", "Uptime"]
    _write_header_row(ws, headers)
    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        rd = device.get("collector_data", {}).get("routing_detail", {}).get("parsed", {})
        site = _derive_site(hostname)
        for bp in rd.get("bgp_summary", []):
            _write_row(ws, row, [
                hostname, site,
                bp.get("neighbor", ""),
                bp.get("asn", ""),
                bp.get("state", ""),
                bp.get("prefixes_received", ""),
                bp.get("up_down", ""),
            ])
            row += 1
    _finalize(ws, "BGPPeers", len(headers), row - 1)


def _build_nat_overview_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="NAT Overview")
    headers = ["Device", "Site/Location", "Inside Intfs", "Outside Intfs",
               "Active Translations", "Peak", "Pool Name",
               "Total Addrs", "Allocated", "Util %"]
    _write_header_row(ws, headers)
    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
        nat_stats = es.get("nat_statistics", {})
        if not nat_stats:
            continue
        site = _derive_site(hostname)
        inside = ", ".join(nat_stats.get("inside_interfaces", []))
        outside = ", ".join(nat_stats.get("outside_interfaces", []))
        pools = nat_stats.get("pools", [])
        if pools:
            for pool in pools:
                _write_row(ws, row, [
                    hostname, site, inside, outside,
                    nat_stats.get("active_translations", ""),
                    nat_stats.get("peak_translations", ""),
                    pool.get("name", ""),
                    pool.get("total_addresses", ""),
                    pool.get("allocated", ""),
                    pool.get("utilization_pct", ""),
                ])
                row += 1
        else:
            _write_row(ws, row, [
                hostname, site, inside, outside,
                nat_stats.get("active_translations", ""),
                nat_stats.get("peak_translations", ""),
                "", "", "", "",
            ])
            row += 1
    _finalize(ws, "NATOverview", len(headers), row - 1)


def _build_acl_audit_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="ACL Audit")
    headers = ["Device", "Site/Location", "ACL Name", "Type",
               "Rule Count", "Zero-Hit Rules", "Has Permit-Any"]
    _write_header_row(ws, headers)
    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
        site = _derive_site(hostname)
        for acl in es.get("access_lists", []):
            entries = acl.get("entries", [])
            zero_hits = sum(1 for e in entries if e.get("hit_count", 0) == 0)
            has_permit_any = any(
                e.get("action", "").lower() == "permit"
                and e.get("protocol", "").lower() == "ip"
                and e.get("source", "").lower() == "any"
                and e.get("destination", "").lower() == "any"
                for e in entries
            )
            _write_row(ws, row, [
                hostname, site,
                acl.get("name", ""),
                acl.get("type", ""),
                len(entries),
                zero_hits,
                "Yes" if has_permit_any else "No",
            ])
            row += 1
    _finalize(ws, "ACLAudit", len(headers), row - 1)


def _build_fhrp_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="FHRP Status")
    headers = ["Device", "Site/Location", "Interface", "Group",
               "Priority", "State", "Virtual IP"]
    _write_header_row(ws, headers)
    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        hsrp = device.get("collector_data", {}).get("hsrp", {}).get("parsed", {})
        site = _derive_site(hostname)
        for e in hsrp.get("entries", []):
            _write_row(ws, row, [
                hostname, site,
                e.get("interface", ""),
                e.get("group", ""),
                e.get("priority", ""),
                e.get("state", ""),
                e.get("virtual_ip", ""),
            ])
            row += 1
    _finalize(ws, "FHRPStatus", len(headers), row - 1)


def _build_edge_interfaces_sheet(wb, inventory_data):
    ws = wb.create_sheet(title="Edge Interfaces")
    headers = ["Device", "Site/Location", "Interface", "IP",
               "Description", "ACL In", "ACL Out", "Proxy ARP", "uRPF"]
    _write_header_row(ws, headers)
    row = 2
    for hostname, device in sorted(inventory_data.get("devices", {}).items()):
        es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
        site = _derive_site(hostname)
        for iface in es.get("ip_interfaces", []):
            _write_row(ws, row, [
                hostname, site,
                iface.get("interface", ""),
                iface.get("ip_address", ""),
                "",
                iface.get("acl_in", ""),
                iface.get("acl_out", ""),
                "Yes" if iface.get("proxy_arp") else "No",
                "Yes" if iface.get("urpf") else "No",
            ])
            row += 1
    _finalize(ws, "EdgeInterfaces", len(headers), row - 1)


def _build_findings_sheet(wb, analysis):
    ws = wb.create_sheet(title="Findings")
    headers = ["Device", "Site/Location", "Severity", "Finding", "Details"]
    _write_header_row(ws, headers)
    row = 2
    for finding in analysis["findings"]:
        hostname = finding.get("hostname", "")
        severity = finding.get("severity", "Info")
        _write_row(ws, row, [
            hostname,
            _derive_site(hostname),
            severity,
            finding["title"],
            finding["description"],
        ])
        if severity == "Critical":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _CRITICAL_FONT
                cell.fill = _CRITICAL_FILL
        elif severity == "Warning":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _WARNING_FONT
                cell.fill = _WARNING_FILL
        row += 1
    if not analysis["findings"]:
        ws.cell(row=2, column=1, value="No findings detected.")
        row = 3
    _finalize(ws, "EdgeFindings", len(headers), row - 1)


# ---------------------------------------------------------------------------
# Report class
# ---------------------------------------------------------------------------

class InternetEdgeReport(BaseReport):
    name = "internet_edge"
    label = "Internet Edge"
    description = "BGP peering, NAT overview, ACL audit, FHRP status, edge interface security"
    category = "Compliance & Config"
    required_collectors = ["routing_detail"]
    supported_formats = ["xlsx", "json", "csv", "xml"]

    def generate(self, inventory_data: dict, fmt: str = "xlsx") -> bytes:
        if fmt != "xlsx":
            return self._generate_for_format(inventory_data, fmt)

        analysis = _analyze_edge(inventory_data)
        wb = Workbook()
        wb.remove(wb.active)

        _build_summary_sheet(wb, analysis, inventory_data)

        if _has_collector_data(inventory_data, "routing_detail"):
            _build_bgp_peers_sheet(wb, inventory_data)

        if _has_collector_data(inventory_data, "edge_services"):
            _build_nat_overview_sheet(wb, inventory_data)
            _build_acl_audit_sheet(wb, inventory_data)

        if _has_collector_data(inventory_data, "hsrp"):
            _build_fhrp_sheet(wb, inventory_data)

        if _has_collector_data(inventory_data, "edge_services"):
            _build_edge_interfaces_sheet(wb, inventory_data)

        _build_findings_sheet(wb, analysis)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_tabular_data(self, inventory_data):
        analysis = _analyze_edge(inventory_data)
        devices = inventory_data.get("devices", {})
        sheets = OrderedDict()

        # Summary
        edge_count = 0
        bgp_peer_count = 0
        nat_active = 0
        for device in devices.values():
            cd = device.get("collector_data", {})
            rd = cd.get("routing_detail", {}).get("parsed", {})
            es = cd.get("edge_services", {}).get("parsed", {})
            bgp = rd.get("bgp_summary", [])
            if bgp or es.get("nat_statistics") or es.get("access_lists"):
                edge_count += 1
            bgp_peer_count += len(bgp)
            nat_active += es.get("nat_statistics", {}).get("active_translations", 0)

        critical = sum(1 for f in analysis["findings"] if f.get("severity") == "Critical")
        warning = sum(1 for f in analysis["findings"] if f.get("severity") == "Warning")
        info = sum(1 for f in analysis["findings"] if f.get("severity") == "Info")

        sheets["Summary"] = (
            ["Metric", "Value"],
            [
                ["Total Devices", len(devices)],
                ["Edge Devices", edge_count],
                ["BGP Peers", bgp_peer_count],
                ["Active NAT Translations", nat_active],
                ["Critical Findings", critical],
                ["Warning Findings", warning],
                ["Info Findings", info],
            ],
        )

        # BGP Peers
        if _has_collector_data(inventory_data, "routing_detail"):
            bgp_headers = ["Device", "Site/Location", "Neighbor", "ASN",
                           "State", "Prefixes Rcvd", "Uptime"]
            bgp_rows = []
            for hostname, device in sorted(devices.items()):
                rd = device.get("collector_data", {}).get("routing_detail", {}).get("parsed", {})
                site = _derive_site(hostname)
                for bp in rd.get("bgp_summary", []):
                    bgp_rows.append([
                        hostname, site, bp.get("neighbor", ""),
                        bp.get("asn", ""), bp.get("state", ""),
                        bp.get("prefixes_received", ""), bp.get("up_down", ""),
                    ])
            sheets["BGP Peers"] = (bgp_headers, bgp_rows)

        # NAT Overview
        if _has_collector_data(inventory_data, "edge_services"):
            nat_headers = ["Device", "Site/Location", "Inside Intfs", "Outside Intfs",
                           "Active Translations", "Peak", "Pool Name",
                           "Total Addrs", "Allocated", "Util %"]
            nat_rows = []
            for hostname, device in sorted(devices.items()):
                es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
                nat_stats = es.get("nat_statistics", {})
                if not nat_stats:
                    continue
                site = _derive_site(hostname)
                inside = ", ".join(nat_stats.get("inside_interfaces", []))
                outside = ", ".join(nat_stats.get("outside_interfaces", []))
                for pool in nat_stats.get("pools", [{}]):
                    nat_rows.append([
                        hostname, site, inside, outside,
                        nat_stats.get("active_translations", ""),
                        nat_stats.get("peak_translations", ""),
                        pool.get("name", ""),
                        pool.get("total_addresses", ""),
                        pool.get("allocated", ""),
                        pool.get("utilization_pct", ""),
                    ])
            sheets["NAT Overview"] = (nat_headers, nat_rows)

            # ACL Audit
            acl_headers = ["Device", "Site/Location", "ACL Name", "Type",
                           "Rule Count", "Zero-Hit Rules", "Has Permit-Any"]
            acl_rows = []
            for hostname, device in sorted(devices.items()):
                es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
                site = _derive_site(hostname)
                for acl in es.get("access_lists", []):
                    entries = acl.get("entries", [])
                    zero_hits = sum(1 for e in entries if e.get("hit_count", 0) == 0)
                    has_permit_any = any(
                        e.get("action", "").lower() == "permit"
                        and e.get("protocol", "").lower() == "ip"
                        and e.get("source", "").lower() == "any"
                        and e.get("destination", "").lower() == "any"
                        for e in entries
                    )
                    acl_rows.append([
                        hostname, site, acl.get("name", ""),
                        acl.get("type", ""), len(entries), zero_hits,
                        "Yes" if has_permit_any else "No",
                    ])
            sheets["ACL Audit"] = (acl_headers, acl_rows)

        # FHRP Status
        if _has_collector_data(inventory_data, "hsrp"):
            fhrp_headers = ["Device", "Site/Location", "Interface", "Group",
                            "Priority", "State", "Virtual IP"]
            fhrp_rows = []
            for hostname, device in sorted(devices.items()):
                hsrp = device.get("collector_data", {}).get("hsrp", {}).get("parsed", {})
                site = _derive_site(hostname)
                for e in hsrp.get("entries", []):
                    fhrp_rows.append([
                        hostname, site, e.get("interface", ""),
                        e.get("group", ""), e.get("priority", ""),
                        e.get("state", ""), e.get("virtual_ip", ""),
                    ])
            sheets["FHRP Status"] = (fhrp_headers, fhrp_rows)

        # Edge Interfaces
        if _has_collector_data(inventory_data, "edge_services"):
            ei_headers = ["Device", "Site/Location", "Interface", "IP",
                          "ACL In", "ACL Out", "Proxy ARP", "uRPF"]
            ei_rows = []
            for hostname, device in sorted(devices.items()):
                es = device.get("collector_data", {}).get("edge_services", {}).get("parsed", {})
                site = _derive_site(hostname)
                for iface in es.get("ip_interfaces", []):
                    ei_rows.append([
                        hostname, site, iface.get("interface", ""),
                        iface.get("ip_address", ""),
                        iface.get("acl_in", ""), iface.get("acl_out", ""),
                        "Yes" if iface.get("proxy_arp") else "No",
                        "Yes" if iface.get("urpf") else "No",
                    ])
            sheets["Edge Interfaces"] = (ei_headers, ei_rows)

        # Findings
        findings_headers = ["Device", "Site/Location", "Severity", "Finding", "Details"]
        findings_rows = []
        for f in analysis["findings"]:
            hostname = f.get("hostname", "")
            findings_rows.append([
                hostname, _derive_site(hostname),
                f.get("severity", "Info"), f["title"], f["description"],
            ])
        sheets["Findings"] = (findings_headers, findings_rows)

        return sheets
