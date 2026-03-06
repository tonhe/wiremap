import io
from openpyxl import load_workbook
from app.reports.routing_summary import RoutingSummaryReport


def test_report_attrs():
    r = RoutingSummaryReport()
    assert r.name == "routing_summary"
    assert "l3_routing" in r.required_collectors


def test_generate_xlsx(sample_inventory):
    r = RoutingSummaryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Protocol Neighbors" in wb.sheetnames
    assert "Routes" in wb.sheetnames
    assert "Routed Interfaces" in wb.sheetnames
    assert "Findings" in wb.sheetnames

    # Protocol Neighbors: Device, Site/Location, Protocol, ...
    ws = wb["Protocol Neighbors"]
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=3).value == "ospf"


def test_generate_xlsx_with_routing_detail(sample_inventory):
    """Verify Route Summary and OSPF Topology tabs appear when routing_detail has data."""
    # Add some routing_detail data
    dev = sample_inventory["devices"]["SW1"]["collector_data"]["routing_detail"]
    dev["parsed"] = {
        "route_summary": [
            {"source": "connected", "count": 5},
            {"source": "ospf 1", "count": 10},
        ],
        "ospf_processes": [
            {"process_id": "1", "router_id": "10.0.0.1", "areas": ["0"]},
        ],
        "ospf_interfaces": [
            {"interface": "Gi0/1", "area": "0", "cost": "1", "state": "DR", "neighbors": "1"},
        ],
        "bgp_summary": [],
        "eigrp_topology": [],
        "eigrp_neighbors": [],
    }

    r = RoutingSummaryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Route Summary" in wb.sheetnames
    assert "OSPF Topology" in wb.sheetnames

    ws = wb["Route Summary"]
    assert ws.cell(row=2, column=1).value == "SW1"

    ws2 = wb["OSPF Topology"]
    assert ws2.cell(row=2, column=3).value == "1"  # Process ID
