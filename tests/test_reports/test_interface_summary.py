import io
from openpyxl import load_workbook
from app.reports.interface_summary import InterfaceSummaryReport


def test_report_attrs():
    r = InterfaceSummaryReport()
    assert r.name == "interface_summary"
    assert "interfaces" in r.required_collectors


def test_generate_xlsx(sample_inventory):
    r = InterfaceSummaryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Interface Status" in wb.sheetnames
    assert "IP Interfaces" in wb.sheetnames
    ws = wb["Interface Status"]
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=2).value == "Gi0/1"
