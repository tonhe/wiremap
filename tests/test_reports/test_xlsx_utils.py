import pytest
from openpyxl import Workbook
from app.reports.xlsx_utils import (
    style_header_row, write_data_rows, auto_width,
    freeze_header, create_sheet,
    HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
)


def test_style_header_row():
    wb = Workbook()
    ws = wb.active
    headers = ["Hostname", "IP", "Type"]
    style_header_row(ws, headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        assert cell.value == header
        assert cell.font == HEADER_FONT
        assert cell.fill == HEADER_FILL
        assert cell.alignment == HEADER_ALIGN


def test_write_data_rows():
    wb = Workbook()
    ws = wb.active
    rows = [["SW1", "10.0.0.1", "cisco_ios"], ["SW2", "10.0.0.2", "arista_eos"]]
    write_data_rows(ws, rows, start_row=2)
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=3, column=2).value == "10.0.0.2"


def test_auto_width():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Short")
    ws.cell(row=2, column=1, value="A much longer cell value here")
    auto_width(ws)
    # Column should be wider than minimum
    assert ws.column_dimensions["A"].width >= 8


def test_freeze_header():
    wb = Workbook()
    ws = wb.active
    freeze_header(ws)
    assert ws.freeze_panes == "A2"


def test_create_sheet():
    wb = Workbook()
    headers = ["Name", "Value"]
    rows = [["a", 1], ["b", 2], ["c", 3]]
    create_sheet(wb, "Test", headers, rows)
    ws = wb["Test"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "a"
    assert ws.cell(row=4, column=2).value == 3
    assert ws.freeze_panes == "A2"
