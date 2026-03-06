import io
from openpyxl import load_workbook
from app.reports.link_inventory import LinkInventoryReport


def test_report_attrs():
    r = LinkInventoryReport()
    assert r.name == "link_inventory"
    assert "cdp_lldp" in r.required_collectors


def test_generate_xlsx(sample_inventory):
    r = LinkInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Links"]
    assert ws.cell(row=1, column=1).value == "Local Device"
    # Should have 1 unique link (SW1-SW2, deduplicated)
    assert ws.cell(row=2, column=1).value is not None
    assert ws.cell(row=3, column=1).value is None  # no second link row


def test_by_device_tab(sample_inventory):
    r = LinkInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "By Device" in wb.sheetnames
