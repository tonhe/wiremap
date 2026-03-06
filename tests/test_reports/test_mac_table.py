import io
from openpyxl import load_workbook
from app.reports.mac_table import MacTableReport


def test_report_attrs():
    r = MacTableReport()
    assert r.name == "mac_table"
    assert "mac_table" in r.required_collectors


def test_generate_xlsx(sample_inventory):
    r = MacTableReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["MAC Table"]
    assert ws.cell(row=1, column=1).value == "Device"
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=2).value == "aabb.cc00.0002"
