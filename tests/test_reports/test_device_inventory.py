import io
import pytest
from openpyxl import load_workbook
from app.reports.device_inventory import DeviceInventoryReport


def test_report_attrs():
    r = DeviceInventoryReport()
    assert r.name == "device_inventory"
    assert "device_inventory" in r.required_collectors


def test_generate_xlsx_devices_sheet(sample_inventory):
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Devices"]
    # Header includes new Platform Description column
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert "Platform Description" in headers
    assert "Software Version" in headers
    # SW1 data
    assert ws.cell(row=2, column=1).value == "SW1"
    # Software version should be populated
    assert ws.cell(row=2, column=7).value == "15.2(4)E10"
    # Serial should be populated
    assert ws.cell(row=2, column=8).value == "FOC1234ABCD"


def test_devices_sheet_platform_description(sample_inventory):
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Devices"]
    # SW1 platform is WS-C3750X-48 -> should match description
    desc_col = None
    for c in range(1, 10):
        if ws.cell(row=1, column=c).value == "Platform Description":
            desc_col = c
            break
    assert desc_col is not None
    desc = ws.cell(row=2, column=desc_col).value
    assert desc and "3750" in desc


def test_generate_xlsx_stack_members_sheet(sample_inventory):
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Stack Members" in wb.sheetnames
    ws = wb["Stack Members"]
    # Headers
    assert ws.cell(row=1, column=1).value == "Hostname"
    assert ws.cell(row=1, column=3).value == "Member #"
    assert ws.cell(row=1, column=4).value == "Role"
    # SW1 has 2 stack members
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=3).value == "1"
    assert ws.cell(row=2, column=4).value == "Active"
    assert ws.cell(row=3, column=1).value == "SW1"
    assert ws.cell(row=3, column=3).value == "2"
    assert ws.cell(row=3, column=4).value == "Standby"
    # SW2 has no stack members -- should not appear
    hostnames = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "SW2" not in hostnames


def test_stack_members_sheet_empty_when_no_stacks(sample_inventory):
    for dev in sample_inventory["devices"].values():
        dev["collector_data"]["device_inventory"]["parsed"]["stack_members"] = []
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Stack Members"]
    assert ws.max_row == 1  # header only


def test_generate_xlsx_modules_sheet(sample_inventory):
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Modules" in wb.sheetnames
    ws = wb["Modules"]
    # Headers
    assert ws.cell(row=1, column=1).value == "Hostname"
    assert ws.cell(row=1, column=3).value == "Slot"
    assert ws.cell(row=1, column=4).value == "Type"
    # SW1 has 2 modules
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=3).value == "1"
    assert ws.cell(row=2, column=4).value == "Supervisor Module"
    assert ws.cell(row=3, column=3).value == "2"
    assert ws.cell(row=3, column=4).value == "Line Card"


def test_modules_sheet_uses_inventory_fallback(sample_inventory):
    for dev in sample_inventory["devices"].values():
        dev["collector_data"]["device_inventory"]["parsed"]["modules"] = []
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Modules"]
    assert ws is not None


def test_can_generate(sample_inventory):
    r = DeviceInventoryReport()
    assert r.can_generate(sample_inventory) is True
    assert r.can_generate({"devices": {}}) is False


def test_three_sheets_present(sample_inventory):
    r = DeviceInventoryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Devices", "Stack Members", "Modules"]
