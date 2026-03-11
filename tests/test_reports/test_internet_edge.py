import io
import json
from openpyxl import load_workbook
from app.reports.internet_edge import InternetEdgeReport


def test_report_attrs():
    r = InternetEdgeReport()
    assert r.name == "internet_edge"
    assert r.label == "Internet Edge"
    assert "routing_detail" in r.required_collectors
    assert "xlsx" in r.supported_formats
    assert "json" in r.supported_formats


def test_generate_xlsx(sample_inventory):
    r = InternetEdgeReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    sheet_names = wb.sheetnames
    assert "Summary" in sheet_names
    assert "Findings" in sheet_names


def test_bgp_peers_tab(sample_inventory):
    r = InternetEdgeReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "BGP Peers" in wb.sheetnames
    ws = wb["BGP Peers"]
    assert ws.cell(row=1, column=1).value == "Device"
    # SW1 has one BGP peer
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=2, column=4).value == "65002"


def test_nat_overview_tab(sample_inventory):
    r = InternetEdgeReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "NAT Overview" in wb.sheetnames
    ws = wb["NAT Overview"]
    assert ws.cell(row=1, column=1).value == "Device"


def test_acl_audit_tab(sample_inventory):
    r = InternetEdgeReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "ACL Audit" in wb.sheetnames
    ws = wb["ACL Audit"]
    assert ws.cell(row=2, column=3).value == "OUTSIDE_IN"


def test_findings_single_homed_bgp(sample_inventory):
    """SW1 has only 1 BGP peer -- should flag single-homed."""
    r = InternetEdgeReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Findings"]
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value and "Single-homed" in str(ws.cell(row=row, column=4).value):
            found = True
            break
    assert found, "Expected single-homed BGP finding"


def test_generate_json(sample_inventory):
    r = InternetEdgeReport()
    data = r.generate(sample_inventory, fmt="json")
    parsed = json.loads(data)
    assert "Summary" in parsed
    assert "Findings" in parsed
