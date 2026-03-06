import io
from openpyxl import load_workbook
from app.reports.arp_summary import ArpSummaryReport


def test_report_attrs():
    r = ArpSummaryReport()
    assert r.name == "arp_summary"
    assert "arp" in r.required_collectors


def test_generate_xlsx(sample_inventory):
    r = ArpSummaryReport()
    data = r.generate(sample_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Detail" in wb.sheetnames
    ws = wb["Detail"]
    assert ws.cell(row=1, column=1).value == "Device"
    # SW1 has 2 ARP entries
    assert ws.cell(row=2, column=1).value == "SW1"
    assert ws.cell(row=3, column=1).value == "SW1"
