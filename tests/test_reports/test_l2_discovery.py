"""Tests for L2 Discovery report -- VLAN documentation format."""
import io
import pytest
from openpyxl import load_workbook
from app.reports.l2_discovery import (
    L2DiscoveryReport, _analyze_vlans, _is_default_name,
    _names_are_similar, _extract_ip_interfaces,
    _has_collector_data, _derive_site,
)


@pytest.fixture
def l2_inventory():
    """Inventory with STP/VLAN + interfaces + stp_detail + switchport data."""
    return {
        "discovery_id": "test",
        "seed_ip": "10.0.0.1",
        "timestamp": "2026-03-05T10:00:00Z",
        "params": {},
        "devices": {
            "CORE-SW-01": {
                "hostname": "CORE-SW-01",
                "mgmt_ip": "10.0.0.1",
                "device_type": "cisco_ios",
                "device_category": "switch",
                "platform": "",
                "collector_data": {
                    "stp_vlan": {
                        "raw": {},
                        "parsed": {
                            "spanning_tree": [],
                            "spanning_tree_root": [
                                {"vlan_id": "10", "root_address": "aabb.cc00.1010",
                                 "root_priority": "4106", "root_cost": "0",
                                 "root_port": "", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "20", "root_address": "aabb.cc00.2020",
                                 "root_priority": "32788", "root_cost": "4",
                                 "root_port": "Gi0/1", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "30", "root_address": "aabb.cc00.3030",
                                 "root_priority": "32798", "root_cost": "0",
                                 "root_port": "", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                            ],
                            "blocked_ports": [
                                {"vlan_id": "20", "interface": "Gi0/2",
                                 "name": "USERS", "status": "BLK",
                                 "role": "Altn", "reason": ""},
                            ],
                            "vlans": [
                                {"vlan_id": "10", "name": "MGMT"},
                                {"vlan_id": "20", "name": "WBC_USERS"},
                                {"vlan_id": "30", "name": "SERVERS"},
                            ],
                            "vtp_status": [],
                        },
                    },
                    "cdp_lldp": {"raw": {}, "parsed": {"neighbors": [
                        {"remote_device": "DIST-SW-01", "local_intf": "Gi0/1",
                         "remote_intf": "Gi0/24", "remote_ip": "10.0.0.2",
                         "protocols": ["CDP"]},
                    ]}},
                    "hsrp": {"raw": {}, "parsed": {"entries": [
                        {"interface": "Vlan10", "group": "0",
                         "priority": "110", "virtual_ip": "10.0.10.254",
                         "state": "Active"},
                        {"interface": "Vlan20", "group": "0",
                         "priority": "110", "virtual_ip": "10.0.20.254",
                         "state": "Active"},
                    ]}},
                    "interfaces": {
                        "raw": {},
                        "parsed": {
                            "interfaces_status": [],
                            "interfaces_description": [
                                {"interface": "Gi0/1", "description": "Uplink to DIST-SW-01"},
                                {"interface": "Loopback0", "description": "Management Loopback"},
                            ],
                            "ip_interfaces": [],
                            "ip_interfaces_full": [
                                {"interface": "Vlan10",
                                 "ip_address": ["10.0.10.1"], "prefix_length": ["24"]},
                                {"interface": "Vlan20",
                                 "ip_address": ["10.0.20.1", "10.0.21.1"],
                                 "prefix_length": ["24", "24"]},
                                {"interface": "Loopback0",
                                 "ip_address": ["10.255.1.1"], "prefix_length": ["32"]},
                                {"interface": "Port-channel1",
                                 "ip_address": ["10.1.255.1"], "prefix_length": ["30"]},
                            ],
                            "etherchannel": [],
                        },
                    },
                    "stp_detail": {
                        "raw": {},
                        "parsed": {
                            "stp_detail": [
                                {"vlan": "10", "interface": "Gi0/1", "role": "Desg",
                                 "state": "FWD", "cost": "4", "topology_changes": "2"},
                                {"vlan": "20", "interface": "Gi0/2", "role": "Root",
                                 "state": "FWD", "cost": "4", "topology_changes": "15"},
                            ],
                            "inconsistent_ports": [],
                            "stp_root_summary": [],
                        },
                    },
                    "switchport": {
                        "raw": {},
                        "parsed": {
                            "switchports": [
                                {"interface": "Gi0/1", "mode": "trunk",
                                 "native_vlan": "1", "allowed_vlans": "ALL"},
                                {"interface": "Gi0/3", "mode": "access",
                                 "access_vlan": "10", "voice_vlan": "100"},
                            ],
                            "port_security": [],
                            "port_security_addresses": [],
                            "errdisable_recovery": [],
                            "storm_control": [],
                        },
                    },
                    "config": {
                        "raw": {},
                        "parsed": {
                            "config": "hostname CORE-SW-01\n!\nspanning-tree portfast bpduguard default\n!",
                        },
                    },
                },
            },
            "DIST-SW-01": {
                "hostname": "DIST-SW-01",
                "mgmt_ip": "10.0.0.2",
                "device_type": "cisco_ios",
                "device_category": "switch",
                "platform": "",
                "collector_data": {
                    "stp_vlan": {
                        "raw": {},
                        "parsed": {
                            "spanning_tree": [],
                            "spanning_tree_root": [
                                {"vlan_id": "10", "root_address": "aabb.cc00.1010",
                                 "root_priority": "4106", "root_cost": "4",
                                 "root_port": "Gi0/1", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "20", "root_address": "aabb.cc00.2020",
                                 "root_priority": "32788", "root_cost": "0",
                                 "root_port": "", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "30", "root_address": "ddee.ff00.3030",
                                 "root_priority": "32798", "root_cost": "0",
                                 "root_port": "", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                            ],
                            "blocked_ports": [
                                {"vlan_id": "10", "interface": "Po50",
                                 "name": "MGMT", "status": "BKN",
                                 "role": "Desg", "reason": "BA_Inc"},
                            ],
                            "vlans": [
                                {"vlan_id": "10", "name": "MGMT"},
                                {"vlan_id": "20", "name": "WBC_USER"},
                                {"vlan_id": "30", "name": "SERVERS"},
                            ],
                            "vtp_status": [],
                        },
                    },
                    "cdp_lldp": {"raw": {}, "parsed": {"neighbors": [
                        {"remote_device": "CORE-SW-01", "local_intf": "Gi0/24",
                         "remote_intf": "Gi0/1", "remote_ip": "10.0.0.1",
                         "protocols": ["CDP"]},
                    ]}},
                    "hsrp": {"raw": {}, "parsed": {"entries": [
                        {"interface": "Vlan10", "group": "0",
                         "priority": "100", "virtual_ip": "10.0.10.254",
                         "state": "Standby"},
                        {"interface": "Vlan20", "group": "0",
                         "priority": "100", "virtual_ip": "10.0.20.254",
                         "state": "Standby"},
                    ]}},
                    "interfaces": {
                        "raw": {},
                        "parsed": {
                            "interfaces_status": [],
                            "interfaces_description": [],
                            "ip_interfaces": [],
                            "ip_interfaces_full": [
                                {"interface": "Vlan10",
                                 "ip_address": ["10.0.10.2"], "prefix_length": ["24"]},
                                {"interface": "Vlan20",
                                 "ip_address": ["10.0.20.2"], "prefix_length": ["24"]},
                                {"interface": "Vlan30",
                                 "ip_address": ["172.16.0.1"], "prefix_length": ["16"]},
                            ],
                            "etherchannel": [],
                        },
                    },
                    "stp_detail": {
                        "raw": {},
                        "parsed": {
                            "stp_detail": [
                                {"vlan": "10", "interface": "Gi0/1", "role": "Root",
                                 "state": "FWD", "cost": "4", "topology_changes": "1"},
                            ],
                            "inconsistent_ports": [],
                            "stp_root_summary": [],
                        },
                    },
                    "switchport": {
                        "raw": {},
                        "parsed": {
                            "switchports": [
                                {"interface": "Gi0/24", "mode": "trunk",
                                 "native_vlan": "1", "allowed_vlans": "10,20,30"},
                                {"interface": "Gi0/5", "mode": "access",
                                 "access_vlan": "20", "voice_vlan": ""},
                            ],
                            "port_security": [],
                            "port_security_addresses": [],
                            "errdisable_recovery": [],
                            "storm_control": [
                                {"interface": "Gi0/5", "level": "10.00"},
                            ],
                        },
                    },
                    "config": {
                        "raw": {},
                        "parsed": {
                            "config": "hostname DIST-SW-01\n!",
                        },
                    },
                },
            },
            "ACCESS-SW-01": {
                "hostname": "ACCESS-SW-01",
                "mgmt_ip": "10.0.0.3",
                "device_type": "cisco_ios",
                "device_category": "switch",
                "platform": "",
                "collector_data": {
                    "stp_vlan": {
                        "raw": {},
                        "parsed": {
                            "spanning_tree": [],
                            "spanning_tree_root": [
                                {"vlan_id": "10", "root_address": "aabb.cc00.1010",
                                 "root_priority": "4106", "root_cost": "8",
                                 "root_port": "Gi0/1", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "20", "root_address": "aabb.cc00.2020",
                                 "root_priority": "32788", "root_cost": "8",
                                 "root_port": "Gi0/1", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                                {"vlan_id": "30", "root_address": "aabb.cc00.3030",
                                 "root_priority": "32798", "root_cost": "4",
                                 "root_port": "Gi0/1", "hello_time": "2",
                                 "max_age": "20", "fwd_delay": "15"},
                            ],
                            "blocked_ports": [
                                {"vlan_id": "10", "interface": "Gi0/2",
                                 "name": "MGMT", "status": "BLK"},
                            ],
                            "vlans": [
                                {"vlan_id": "10", "name": "MGMT"},
                                {"vlan_id": "20", "name": "WIRELESS_NETWORK"},
                                {"vlan_id": "30", "name": "SERVERS"},
                            ],
                            "vtp_status": [],
                        },
                    },
                    "cdp_lldp": {"raw": {}, "parsed": {"neighbors": []}},
                    "hsrp": {"raw": {}, "parsed": {"entries": []}},
                    "interfaces": {
                        "raw": {},
                        "parsed": {
                            "interfaces_status": [],
                            "interfaces_description": [],
                            "ip_interfaces": [],
                            "ip_interfaces_full": [],
                            "etherchannel": [],
                        },
                    },
                    "stp_detail": {
                        "raw": {},
                        "parsed": {
                            "stp_detail": [],
                            "inconsistent_ports": [],
                            "stp_root_summary": [],
                        },
                    },
                    "switchport": {
                        "raw": {},
                        "parsed": {
                            "switchports": [],
                            "port_security": [],
                            "port_security_addresses": [],
                            "errdisable_recovery": [],
                            "storm_control": [],
                        },
                    },
                    "config": {
                        "raw": {},
                        "parsed": {
                            "config": "hostname ACCESS-SW-01\n!",
                        },
                    },
                },
            },
        },
    }


# --- Unit tests for helpers ---

def test_is_default_name():
    assert _is_default_name("default") is True
    assert _is_default_name("VLAN0010") is True
    assert _is_default_name("VLAN10") is True
    assert _is_default_name("VLAN0001") is True
    assert _is_default_name("MGMT") is False
    assert _is_default_name("WBC_USERS") is False


def test_names_are_similar():
    assert _names_are_similar("WBC_USERS", "WBC_USERS") is True
    assert _names_are_similar("WBC_USERS", "WBC_USER") is True
    assert _names_are_similar("WBC_USERS", "WIRELESS_NETWORK") is False
    assert _names_are_similar("VLAN0020", "VLAN20") is True


def test_has_collector_data():
    inv = {
        "devices": {
            "SW1": {
                "collector_data": {
                    "stp_detail": {"parsed": {"stp_detail": [{"vlan": "1"}]}},
                    "switchport": {"parsed": {}},
                }
            }
        }
    }
    assert _has_collector_data(inv, "stp_detail") is True
    assert _has_collector_data(inv, "switchport") is False
    assert _has_collector_data(inv, "nonexistent") is False
    assert _has_collector_data({"devices": {}}, "stp_detail") is False


def test_derive_site():
    assert _derive_site("NYC-CORE-01") == "NYC"
    assert _derive_site("LAX-SW-02") == "LAX"
    assert _derive_site("SWITCH1") == ""
    assert _derive_site("") == ""
    assert _derive_site(None) == ""
    assert _derive_site("A-B") == "A"


# --- IP interface extraction ---

def test_extract_ip_interfaces_ios(l2_inventory):
    device = l2_inventory["devices"]["CORE-SW-01"]
    result = _extract_ip_interfaces(device)
    assert len(result) == 4
    vlan20 = [r for r in result if r["interface"] == "Vlan20"][0]
    assert len(vlan20["ips"]) == 2
    assert vlan20["ips"][0]["secondary"] is False
    assert vlan20["ips"][1]["secondary"] is True


def test_extract_ip_interfaces_nxos():
    device = {
        "collector_data": {
            "interfaces": {
                "parsed": {
                    "ip_interfaces_full": [
                        {"interface": "Vlan1",
                         "primary_ip_address": "10.1.253.253",
                         "primary_ip_subnet": "10.1.253.0/24",
                         "secondary_ip_address": ["10.1.50.254"],
                         "secondary_ip_subnet": ["10.1.50.0/24"]},
                    ],
                },
            },
        },
    }
    result = _extract_ip_interfaces(device)
    assert len(result) == 1
    assert len(result[0]["ips"]) == 2
    assert result[0]["ips"][0]["ip"] == "10.1.253.253"
    assert result[0]["ips"][0]["prefix"] == "24"
    assert result[0]["ips"][1]["secondary"] is True


# --- VLAN analysis ---

def test_analyze_vlans_consistent(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    vlan10 = result["vlans"]["10"]
    assert vlan10["consistent"] is True
    assert vlan10["root_bridge_mac"] == "aabb.cc00.1010"
    assert vlan10["root_device"] == "CORE-SW-01"
    assert vlan10["root_resolved"] is True
    assert vlan10["switch_count"] == 3


def test_analyze_vlans_mismatch(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    vlan30 = result["vlans"]["30"]
    assert vlan30["consistent"] is False
    assert len(vlan30["root_bridges"]) == 2


def test_analyze_vlans_root_resolved(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    assert result["vlans"]["10"]["root_device"] == "CORE-SW-01"
    assert result["vlans"]["20"]["root_device"] == "DIST-SW-01"


def test_analyze_vlans_mac_map(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    mac_map = result["mac_map"]
    assert mac_map["aabb.cc00.1010"] == "CORE-SW-01"
    assert mac_map["aabb.cc00.2020"] == "DIST-SW-01"


def test_analyze_vlans_svi_ips(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    svi_10 = result["svi_ips"]["10"]
    assert len(svi_10) == 2
    svi_20 = result["svi_ips"]["20"]
    assert len(svi_20) == 3
    secondaries = [e for e in svi_20 if e["secondary"]]
    assert len(secondaries) == 1


def test_analyze_vlans_name_discrepancies(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    assert len(result["name_discrepancies"]) > 0
    vlan20_disc = [d for d in result["name_discrepancies"] if d["vlan_id"] == "20"]
    assert len(vlan20_disc) > 0


def test_findings_blocked_ports(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    bkn = [f for f in result["findings"] if "BA_Inc" in f["title"]]
    assert len(bkn) == 1
    assert bkn[0]["hostname"] == "DIST-SW-01"
    assert "BA_Inc" in bkn[0]["description"]


def test_findings_default_stp_priority(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    default_prio = [f for f in result["findings"] if "default STP" in f["title"]]
    assert len(default_prio) >= 1
    vlan20_finding = [f for f in default_prio if "VLAN 20" in f["title"]]
    assert len(vlan20_finding) == 1


def test_findings_large_subnet(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    large = [f for f in result["findings"] if "/16" in f["title"]]
    assert len(large) == 1
    assert "VLAN 30" in large[0]["title"]


def test_analyze_hsrp_vips(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    hsrp_10 = result["hsrp_vips"]["10"]
    assert len(hsrp_10) == 2
    active = [e for e in hsrp_10 if e["state"] == "Active"]
    standby = [e for e in hsrp_10 if e["state"] == "Standby"]
    assert len(active) == 1
    assert len(standby) == 1
    assert active[0]["virtual_ip"] == "10.0.10.254"


def test_findings_root_mismatch(l2_inventory):
    result = _analyze_vlans(l2_inventory)
    mismatch = [f for f in result["findings"] if "mismatch" in f["title"]]
    assert len(mismatch) >= 1
    assert "VLAN 30" in mismatch[0]["title"]


# --- Report generation ---

def test_report_attrs():
    r = L2DiscoveryReport()
    assert r.name == "l2_discovery"
    assert "stp_vlan" in r.required_collectors
    assert "cdp_lldp" in r.required_collectors
    assert "interfaces" in r.required_collectors


def test_generate_xlsx_sheets(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "VLAN Documentation" in wb.sheetnames
    assert "Findings" in wb.sheetnames
    assert "Root Bridge IDs" in wb.sheetnames
    assert "STP Topology" in wb.sheetnames
    assert "Port Security" in wb.sheetnames
    assert "Trunk Summary" in wb.sheetnames
    assert len(wb.sheetnames) == 7


def test_generate_summary_tab(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    ws = wb["Summary"]
    assert ws.cell(row=1, column=1).value == "Metric"
    assert ws.cell(row=1, column=2).value == "Value"
    metrics = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            metrics.append(val)
    assert "Device Count" in metrics
    assert "VLAN Count" in metrics
    assert "Critical Findings" in metrics
    assert "Warning Findings" in metrics
    assert "Info Findings" in metrics
    device_row = [r for r in range(2, ws.max_row + 1)
                  if ws.cell(row=r, column=1).value == "Device Count"][0]
    assert ws.cell(row=device_row, column=2).value == 3


def test_generate_vlan_doc_structure(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["VLAN Documentation"]
    assert ws.cell(row=1, column=1).value == "VLAN ID"
    assert ws.cell(row=1, column=2).value == "Site/Location"
    assert ws.cell(row=1, column=3).value == "VLAN Name"
    assert ws.cell(row=1, column=4).value == "Seen On (Switches)"
    assert ws.cell(row=1, column=5).value == "STP Root Bridge"
    assert ws.cell(row=1, column=7).value == "SVI IP / CIDR"
    assert ws.cell(row=1, column=8).value == "HSRP VIP"
    assert ws.cell(row=1, column=9).value == "Blocked Ports"
    assert ws.cell(row=1, column=10).value == "Notes"
    assert ws.cell(row=2, column=1).value == 10


def test_generate_root_bridge_resolved(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["VLAN Documentation"]
    assert ws.cell(row=2, column=5).value == "CORE-SW-01"
    fill = ws.cell(row=2, column=5).fill
    assert fill.start_color.rgb.endswith("E2EFDA")


def test_generate_svi_multiline(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["VLAN Documentation"]
    svi_val = ws.cell(row=3, column=7).value
    assert "10.0.20.1/24" in svi_val
    assert "(secondary)" in svi_val


def test_generate_hsrp_column(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["VLAN Documentation"]
    hsrp_val = ws.cell(row=2, column=8).value
    assert "10.0.10.254" in hsrp_val
    assert "Active" in hsrp_val
    assert "Standby" in hsrp_val


def test_generate_findings(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Findings"]
    assert ws.cell(row=1, column=1).value == "Hostname"
    assert ws.cell(row=1, column=2).value == "Site/Location"
    assert ws.cell(row=1, column=3).value == "Port"
    assert ws.cell(row=1, column=4).value == "Notes"


def test_generate_root_bridge_ids(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Root Bridge IDs"]
    assert ws.cell(row=1, column=1).value == "Hostname"
    assert ws.cell(row=1, column=2).value == "Site/Location"
    assert ws.cell(row=1, column=3).value == "Root Bridge ID"
    hostnames = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            hostnames.append(val)
    assert "CORE-SW-01" in hostnames


def test_generate_stp_topology_tab(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "STP Topology" in wb.sheetnames
    ws = wb["STP Topology"]
    assert ws.cell(row=1, column=1).value == "Device"
    assert ws.cell(row=1, column=2).value == "Site/Location"
    assert ws.cell(row=1, column=3).value == "VLAN"
    assert ws.cell(row=1, column=8).value == "Topology Changes"
    data_rows = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            data_rows += 1
    assert data_rows == 3


def test_generate_port_security_tab(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Port Security" in wb.sheetnames
    ws = wb["Port Security"]
    assert ws.cell(row=1, column=1).value == "Device"
    assert ws.cell(row=1, column=4).value == "Mode"
    assert ws.cell(row=1, column=5).value == "BPDU Guard"
    assert ws.cell(row=1, column=9).value == "Voice VLAN"
    found_core_access = False
    found_dist_access = False
    for row in range(2, ws.max_row + 1):
        device = ws.cell(row=row, column=1).value
        mode = ws.cell(row=row, column=4).value
        bpdu = ws.cell(row=row, column=5).value
        if device == "CORE-SW-01" and mode == "access":
            found_core_access = True
            assert bpdu == "Yes"
        if device == "DIST-SW-01" and mode == "access":
            found_dist_access = True
            assert bpdu == "No"
            # "No" cells should have light red fill
            assert ws.cell(row=row, column=5).fill.start_color.rgb == "00FFC7CE"
    assert found_core_access
    assert found_dist_access


def test_generate_trunk_summary_tab(l2_inventory):
    r = L2DiscoveryReport()
    data = r.generate(l2_inventory)
    wb = load_workbook(io.BytesIO(data))
    assert "Trunk Summary" in wb.sheetnames
    ws = wb["Trunk Summary"]
    assert ws.cell(row=1, column=1).value == "Device"
    assert ws.cell(row=1, column=5).value == "Allowed VLANs"
    assert ws.cell(row=1, column=6).value == "Neighbor"
    found_core_trunk = False
    found_dist_trunk = False
    for row in range(2, ws.max_row + 1):
        device = ws.cell(row=row, column=1).value
        native = ws.cell(row=row, column=4).value
        allowed = ws.cell(row=row, column=5).value
        if device == "CORE-SW-01":
            found_core_trunk = True
            assert allowed == "ALL"
            # Unpruned trunk should have light red fill on allowed VLANs
            assert ws.cell(row=row, column=5).fill.start_color.rgb == "00FFC7CE"
        if device == "DIST-SW-01":
            found_dist_trunk = True
            assert allowed == "10,20,30"
    assert found_core_trunk
    assert found_dist_trunk


def test_conditional_sheets_absent_without_data():
    """STP Topology, Port Security, Trunk Summary not added without data."""
    inv = {
        "devices": {
            "SW1": {
                "hostname": "SW1",
                "collector_data": {
                    "stp_vlan": {"raw": {}, "parsed": {
                        "spanning_tree": [], "spanning_tree_root": [],
                        "blocked_ports": [], "vlans": [], "vtp_status": [],
                    }},
                    "cdp_lldp": {"raw": {}, "parsed": {"neighbors": []}},
                    "hsrp": {"raw": {}, "parsed": {"entries": []}},
                    "interfaces": {"raw": {}, "parsed": {
                        "interfaces_status": [], "interfaces_description": [],
                        "ip_interfaces": [], "ip_interfaces_full": [],
                        "etherchannel": [],
                    }},
                },
            },
        },
    }
    r = L2DiscoveryReport()
    data = r.generate(inv)
    wb = load_workbook(io.BytesIO(data))
    assert "STP Topology" not in wb.sheetnames
    assert "Port Security" not in wb.sheetnames
    assert "Trunk Summary" not in wb.sheetnames
    assert len(wb.sheetnames) == 4
